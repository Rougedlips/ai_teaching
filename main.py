from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Form

from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse

from sqlalchemy.orm import Session, aliased
from pydantic import BaseModel
import models
import schemas
from database import engine, get_db, SessionLocal
import jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta, timezone
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError

from io import BytesIO
from typing import Optional, Any, List

from pathlib import Path
import io
import re
import tokenize

import json
import shutil
import tempfile
import zipfile
import tarfile
import importlib
import time
from http.client import IncompleteRead
import urllib.request
import urllib.error
import xml.etree.ElementTree as ET






try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None
    Workbook = None

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")

SECRET_KEY = "my_super_secret_graduation_key"
ALGORITHM = "HS256"

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="智能教学辅助系统 API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

USERNAME_ALIASES = {
    "techer_demo": "teacher_demo",
}


def _normalize_username(username: Optional[str]) -> str:
    text = (username or "").strip()
    return USERNAME_ALIASES.get(text, text)



def _find_user_by_username(db: Session, username: Optional[str]):
    raw = (username or "").strip()
    if not raw:
        return None

    normalized = _normalize_username(raw)
    user = db.query(models.User).filter(models.User.username == normalized).first()
    if user:
        return user

    if normalized != raw:
        return db.query(models.User).filter(models.User.username == raw).first()

    return None



def _canonical_teacher_id(db: Session, teacher_id: Optional[int]) -> Optional[int]:
    if teacher_id is None:
        return None

    teacher = db.query(models.User).filter(
        models.User.id == teacher_id,
        models.User.role.in_(["teacher", "admin"])
    ).first()
    if not teacher:
        return None

    canonical_user = _find_user_by_username(db, teacher.username)
    if canonical_user and canonical_user.role in ["teacher", "admin"]:
        return canonical_user.id

    return teacher.id



def _require_teacher_user(db: Session, teacher_id: int):
    canonical_id = _canonical_teacher_id(db, teacher_id)
    if canonical_id is None:
        raise HTTPException(status_code=404, detail="老师不存在")

    teacher = db.query(models.User).filter(
        models.User.id == canonical_id,
        models.User.role.in_(["teacher", "admin"])
    ).first()
    if not teacher:
        raise HTTPException(status_code=404, detail="老师不存在")

    return teacher



def _list_distinct_teacher_users(db: Session):
    items = db.query(models.User).filter(models.User.role.in_(["teacher", "admin"])).order_by(models.User.id.desc()).all()
    username_map = {u.username: u for u in items}
    result = []
    seen_ids = set()

    for user in items:
        canonical_user = username_map.get(_normalize_username(user.username), user)
        if canonical_user.id in seen_ids:
            continue
        seen_ids.add(canonical_user.id)
        result.append(canonical_user)

    return result



def _repair_teacher_alias_data():
    db = SessionLocal()
    try:
        for alias_username, canonical_username in USERNAME_ALIASES.items():
            alias_user = db.query(models.User).filter(
                models.User.username == alias_username,
                models.User.role.in_(["teacher", "admin"])
            ).first()
            canonical_user = db.query(models.User).filter(
                models.User.username == canonical_username,
                models.User.role.in_(["teacher", "admin"])
            ).first()

            if not alias_user or not canonical_user or alias_user.id == canonical_user.id:
                continue

            db.query(models.Course).filter(models.Course.teacher_id == alias_user.id).update(
                {models.Course.teacher_id: canonical_user.id},
                synchronize_session=False,
            )
            db.query(models.TeachingClass).filter(models.TeachingClass.teacher_id == alias_user.id).update(
                {models.TeachingClass.teacher_id: canonical_user.id},
                synchronize_session=False,
            )
            db.query(models.ReportTask).filter(models.ReportTask.teacher_id == alias_user.id).update(
                {models.ReportTask.teacher_id: canonical_user.id},
                synchronize_session=False,
            )
            db.query(models.User).filter(models.User.teacher_id == alias_user.id).update(
                {models.User.teacher_id: canonical_user.id},
                synchronize_session=False,
            )
            alias_user.real_name = canonical_user.real_name
            alias_user.role = canonical_user.role
            alias_user.is_admin = canonical_user.is_admin


        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()



_repair_teacher_alias_data()

# 老师请求 AI 评测的数据格式
class AIEvaluateRequest(BaseModel):
    code_content: str
    model_choice: str


class AIProviderConfig(BaseModel):
    base_url: str
    api_key: str
    model_name: str


class AIReviewRequest(BaseModel):
    code_content: str
    prompt_template: str = ""
    assignment_prompt_template: Optional[str] = None
    report_prompt_template: Optional[str] = None
    task_type: str = "assignment"
    functionality_checks: Optional[str] = None
    ai_model: str
    submission_id: Optional[int] = None
    assignment_id: Optional[int] = None
    temperature: float = 0.2
    max_tokens: int = 100000
    provider: AIProviderConfig




# 测试接口
@app.get("/")
def read_root():
    return {"message": "你的 FastAPI 后端正在完美运行！"}

# ================= 改造后的：用户注册接口 =================
@app.post("/users/")
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    normalized_username = _normalize_username(user.username)
    db_user = db.query(models.User).filter(models.User.username == normalized_username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="用户名已被注册，请换一个")

    canonical_teacher_id = _canonical_teacher_id(db, user.teacher_id)
    if user.teacher_id is not None and canonical_teacher_id is None:
        raise HTTPException(status_code=400, detail="teacher_id 无效")

    # 🌟 核心改变：把明文密码变成极其安全的哈希乱码！
    hashed_password = pwd_context.hash(user.password)

    target_role = "admin" if user.is_admin and user.role == "teacher" else user.role
    new_user = models.User(
        username=normalized_username,
        password=hashed_password, # 存入数据库的是乱码，连管理员都看不懂
        role=target_role,
        is_admin=(target_role == "admin"),
        real_name=user.real_name,
        student_no=user.student_no,
        grade=user.grade,
        class_name=user.class_name,
        teacher_id=canonical_teacher_id
    )
    try:
        db.add(new_user)
        db.flush()

        if target_role == "student":
            _ensure_student_class_relation(db, new_user)

        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="账号或学号已存在")

    db.refresh(new_user)

    return {"message": "用户注册成功！", "username": new_user.username}



@app.get("/users/students")
def list_students(db: Session = Depends(get_db)):
    Teacher = aliased(models.User)
    rows = db.query(models.User, Teacher).outerjoin(
        Teacher, models.User.teacher_id == Teacher.id
    ).filter(models.User.role == "student").all()

    return [
        {
            "id": stu.id,
            "username": stu.username,
            "student_no": stu.student_no,
            "real_name": stu.real_name,
            "grade": stu.grade,
            "class_name": stu.class_name,
            "teacher_id": stu.teacher_id,
            "teacher_name": tea.real_name if tea and tea.real_name else (tea.username if tea else None),
        }
        for stu, tea in rows
    ]


@app.get("/users/teachers")
def list_teachers(db: Session = Depends(get_db)):
    teachers = _list_distinct_teacher_users(db)
    return [
        {
            "id": u.id,
            "username": u.username,
            "real_name": u.real_name,
            "role": u.role,
            "is_admin": (u.role == "admin") or bool(getattr(u, "is_admin", False)),
        }
        for u in teachers
    ]

# ================= ⚡ 新增：真正的登录接口 ⚡ =================
@app.post("/login/")
def login(user: schemas.UserLogin, db: Session = Depends(get_db)):
    # 1. 查无此人？
    db_user = _find_user_by_username(db, user.username)
    if not db_user:
        raise HTTPException(status_code=400, detail="用户名或密码错误")

    # 2. 密码不对？（将前端传来的明文，和数据库里的乱码进行神级比对）
    if not pwd_context.verify(user.password, db_user.password):
        raise HTTPException(status_code=400, detail="用户名或密码错误")

    # 3. 登录成功！颁发数字通行证 (JWT Token)
    final_role = "admin" if (db_user.role == "admin" or bool(getattr(db_user, "is_admin", False))) else db_user.role
    expire = datetime.now(timezone.utc) + timedelta(days=1) # 有效期1天
    to_encode = {
        "sub": db_user.username,
        "role": final_role,
        "user_id": db_user.id,
        "exp": expire
    }
    # 盖章生成 Token
    token = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

    return {
        "id": db_user.id,
        "message": "登录成功！",
        "token": token,
        "role": final_role,
        "username": db_user.username,
        "real_name": db_user.real_name,
    }

# ================= 下面是你之前写好的业务接口 =================
@app.post("/courses/")
def create_course(course: schemas.CourseCreate, db: Session = Depends(get_db)):
    canonical_teacher_id = _canonical_teacher_id(db, course.teacher_id)
    if canonical_teacher_id is None:
        raise HTTPException(status_code=400, detail="不合法的老师ID")

    teacher = db.query(models.User).filter(models.User.id == canonical_teacher_id).first()
    if not teacher or teacher.role not in ["teacher", "admin"]:
        raise HTTPException(status_code=400, detail="不合法的老师ID")
    db_course = db.query(models.Course).filter(models.Course.course_code == course.course_code).first()
    if db_course:
        raise HTTPException(status_code=400, detail="课程代码重复")
    new_course = models.Course(course_name=course.course_name, course_code=course.course_code, teacher_id=canonical_teacher_id)
    db.add(new_course)
    db.commit()
    db.refresh(new_course)
    return {"message": "课程创建成功！", "course_id": new_course.id}

# ================= ⚡ 老师发布真实作业接口 ⚡ =================
@app.post("/assignments/")
def create_assignment(assign: schemas.AssignmentCreate, db: Session = Depends(get_db)):
    if assign.course_id is not None:
        target_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == assign.course_id).first()
        if not target_class:
            raise HTTPException(status_code=404, detail="目标班级不存在")

    target_type = _normalize_target_type(assign.target_type)

    new_assign = models.Assignment(
        title=assign.title,
        description=assign.description,
        deadline=assign.deadline,
        course_id=assign.course_id,
        ai_criteria=assign.ai_criteria
    )
    db.add(new_assign)
    db.commit()
    db.refresh(new_assign)

    if assign.course_id is not None:
        db.add(models.AssignmentPublishConfig(
            assignment_id=new_assign.id,
            class_id=assign.course_id,
            target_type=target_type
        ))
        db.commit()

    return {
        "message": "作业发布成功！已实时同步给所有学生。",
        "assignment_id": new_assign.id,
        "target_type": target_type
    }



@app.post("/report-tasks/")
def create_report_task(payload: schemas.ReportTaskCreate, db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    teaching_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == payload.class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="目标班级不存在")

    target_type = _normalize_target_type(payload.target_type)

    task = models.ReportTask(
        title=payload.title,
        description=payload.description,
        deadline=payload.deadline,
        class_id=payload.class_id,
        teacher_id=teaching_class.teacher_id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    db.add(models.ReportTaskPublishConfig(
        report_task_id=task.id,
        class_id=payload.class_id,
        target_type=target_type
    ))
    db.commit()

    return {
        "message": "报告任务发布成功",
        "report_task_id": task.id,
        "target_type": target_type
    }






@app.put("/teachers/{teacher_id}/classes/{class_id}/report-tasks/{report_task_id}")
def update_teacher_class_report_task(
    teacher_id: int,
    class_id: int,
    report_task_id: int,
    payload: schemas.ReportTaskUpdate,
    db: Session = Depends(get_db)
):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    teacher_id = _require_teacher_user(db, teacher_id).id

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在或不属于当前老师")

    task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_task_id,
        models.ReportTask.class_id == class_id,
        models.ReportTask.teacher_id == teacher_id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    if payload.title is not None:
        task.title = payload.title
    if payload.description is not None:
        task.description = payload.description
    if payload.deadline is not None:
        task.deadline = payload.deadline

    db.commit()
    return {"message": "报告任务信息已更新"}


@app.put("/teachers/{teacher_id}/classes/{class_id}/assignments/{assignment_id}")
def update_teacher_class_assignment(


    teacher_id: int,
    class_id: int,
    assignment_id: int,
    payload: schemas.AssignmentUpdate,
    db: Session = Depends(get_db)
):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    teacher_id = _require_teacher_user(db, teacher_id).id

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在或不属于当前老师")

    assignment = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="作业不存在")
    if assignment.course_id != class_id:
        raise HTTPException(status_code=400, detail="该作业不属于当前班级")

    if payload.title is not None:
        assignment.title = payload.title
    if payload.description is not None:
        assignment.description = payload.description
    if payload.deadline is not None:
        assignment.deadline = payload.deadline
    if payload.ai_criteria is not None:
        assignment.ai_criteria = payload.ai_criteria

    db.commit()
    return {"message": "作业信息已更新"}


# ================= ⚡ 学生获取最新作业接口 ⚡ =================
@app.get("/assignments/latest")
def get_latest_assignment(db: Session = Depends(get_db)):
    # 找最新发布的一条作业记录 (按 ID 倒序排)
    assign = db.query(models.Assignment).order_by(models.Assignment.id.desc()).first()
    if not assign:
        return {"title": "暂无作业", "description": "老师还没有发布新作业哦", "deadline": "-"}
    return {"title": assign.title, "description": assign.description, "deadline": assign.deadline}

# ================= 改造 1：学生提交接口（盲交，只存数据库，不调AI） =================
@app.post("/submissions/")
def create_submission(submission: schemas.SubmissionCreate, db: Session = Depends(get_db)):
    new_submission = models.Submission(
        assignment_id=submission.assignment_id, 
        student_id=submission.student_id, 
        code_content=submission.code_content
    )
    db.add(new_submission)
    db.commit()
    db.refresh(new_submission)
    
    # 返回给学生等待老师批阅的提示
    return {"message": "代码已成功提交！请等待老师使用 AI 辅助批阅并生成学情报告。", "submission_id": new_submission.id}

# ================= 改造 2：教师端专属的“多模型评测”接口 (Mock版，兼容保留) =================
@app.post("/ai_evaluate/")
def ai_evaluate(req: AIEvaluateRequest):
    if req.model_choice == "qwen":
        ai_feedback = "【通义千问 评测结果】\n代码逻辑基本正确。但在 V2X 传感器数据融合处缺少异常捕获。建议评分：85分。"
    elif req.model_choice == "deepseek":
        ai_feedback = "【DeepSeek-Coder 评测结果】\n检测到内存泄漏风险！在模型前向传播结束后，未清理 GPU 显存。建议评分：78分。"
    elif req.model_choice == "gpt":
        ai_feedback = "【OpenAI GPT-4 评测结果】\n整体架构符合最佳实践，但在 L2 Error 误差计算时可以使用更高效的矩阵运算。建议评分：92分。"
    else:
        ai_feedback = "未知的模型选择。"

    return {"message": f"已成功调用 {req.model_choice} 模型", "ai_feedback": ai_feedback}


def _build_chat_completions_url(base_url: str) -> str:
    text = (base_url or "").strip().rstrip("/")
    if not text:
        return ""
    if text.endswith("/chat/completions"):
        return text
    if text.endswith("/v1") or text.endswith("/v2") or text.endswith("/v3"):
        return f"{text}/chat/completions"
    return f"{text}/chat/completions"


def _extract_ai_text(payload: Any) -> str:
    if not isinstance(payload, dict):
        return ""

    choices = payload.get("choices")
    if isinstance(choices, list) and choices:
        msg = choices[0].get("message") if isinstance(choices[0], dict) else None
        if isinstance(msg, dict):
            content = msg.get("content")
            if isinstance(content, str):
                return content.strip()

    result = payload.get("result")
    if isinstance(result, str):
        return result.strip()

    output = payload.get("output")
    if isinstance(output, dict):
        text = output.get("text")
        if isinstance(text, str):
            return text.strip()

    return ""


def _decode_escaped_json_string(text: str) -> str:
    try:
        return bytes(text, "utf-8").decode("unicode_escape")
    except Exception:
        return text


def _recover_text_from_truncated_json(raw: str) -> str:
    src = (raw or "").strip()
    if not src:
        return ""

    patterns = [
        r'"content"\s*:\s*"((?:\\.|[^"\\])*)"',
        r'"text"\s*:\s*"((?:\\.|[^"\\])*)"',
        r'"result"\s*:\s*"((?:\\.|[^"\\])*)"'
    ]

    candidates: list[str] = []
    for p in patterns:
        for m in re.finditer(p, src, flags=re.S):
            seg = _decode_escaped_json_string(m.group(1))
            seg = seg.replace("\\n", "\n").strip()
            if seg:
                candidates.append(seg)

    if not candidates:
        return ""

    best = max(candidates, key=len).strip()
    if len(best) < 60:
        return ""

    return best + "\n\n（提示：该结果由中断响应自动恢复，可能存在少量截断）"



def _parse_total_score(ai_feedback: str) -> int:
    text = ai_feedback or ""
    m = re.search(r"本组代码总分为[:：]\s*(\d{1,3})\s*分", text)
    if m:
        return max(0, min(100, int(m.group(1))))

    score_lines = [
        re.search(r"1\.\s*功能评分[:：]\s*(\d{1,3})\s*分", text),
        re.search(r"2\.\s*代码质量[:：]\s*(\d{1,3})\s*分", text),
        re.search(r"3\.\s*原创性分析[:：]\s*(\d{1,3})\s*分", text),
    ]
    vals = [int(x.group(1)) for x in score_lines if x]
    if vals:
        return max(0, min(100, sum(vals)))
    return 0


def _estimate_text_tokens(text: str) -> int:
    if not text:
        return 0

    cjk_count = len(re.findall(r"[\u4e00-\u9fff]", text))
    non_cjk_text = re.sub(r"[\u4e00-\u9fff]", "", text)
    non_cjk_tokens = max(0, (len(non_cjk_text) + 3) // 4)
    return cjk_count + non_cjk_tokens


def _estimate_messages_tokens(messages: list[dict[str, str]]) -> int:
    total = 2
    for msg in messages:
        total += 4
        total += _estimate_text_tokens(str(msg.get("content") or ""))
        total += _estimate_text_tokens(str(msg.get("role") or ""))
    return total



ARCHIVE_MARKER = "[压缩包提交]"
REPORT_FILE_MARKER = "[报告文件]"
REPORT_AI_HISTORY_DIRNAME = "report_ai_history"
REPORT_AI_RECORD_START = "AI_REVIEW_RECORD_START"
REPORT_AI_RECORD_END = "AI_REVIEW_RECORD_END"
MAX_PROMPT_INSTRUCTION_CHARS = 20000


MAX_MODEL_RETRIES = 3
MODEL_RETRY_BACKOFF_SECONDS = 1.0
MIN_RETRY_MAX_TOKENS = 1024
MAX_BACKEND_TOKENS = 100000
RETRY_MAX_TOKENS_RATIO = 0.5
RETRY_PROMPT_SHRINK_RATIO = 0.55
MIN_RETRY_PROMPT_CHARS = 12000
ASSUMED_CONTEXT_WINDOW_TOKENS = 128000
CONTEXT_WINDOW_RESERVED_TOKENS = 2048
DEEPSEEK_CONTEXT_WINDOW_TOKENS = 128000
DEEPSEEK_CHAT_MAX_OUTPUT_TOKENS = 8192
DEEPSEEK_REASONER_MAX_OUTPUT_TOKENS = 64000
DEFAULT_PROMPT_SOFT_LIMIT_TOKENS = 40000
DEEPSEEK_CHAT_PROMPT_SOFT_LIMIT_TOKENS = 36000
DEEPSEEK_REASONER_PROMPT_SOFT_LIMIT_TOKENS = 50000
MIN_CLIPPED_CODE_CHARS = 6000









SOURCE_SUFFIXES = {
    ".py", ".java", ".kt", ".js", ".ts", ".tsx", ".jsx", ".vue", ".go", ".rs", ".cpp", ".cc", ".c", ".h",
    ".hpp", ".cs", ".php", ".rb", ".swift", ".scala", ".sql", ".sh", ".bash", ".zsh", ".yml", ".yaml", ".json",
    ".xml", ".html", ".css", ".scss", ".less", ".md", ".txt", ".ini", ".toml", ".conf", ".env"
}
PRIORITY_SOURCE_SUFFIXES = [
    ".py", ".java", ".vue", ".js", ".ts", ".tsx", ".jsx", ".kt", ".go", ".rs", ".cpp", ".c", ".cs", ".php", ".sql"
]
SKIP_DIR_NAMES = {"__pycache__", ".git", ".idea", ".vscode", "node_modules", "venv", ".venv", "dist", "build", "target"}


def _is_supported_source_file(path: Path) -> bool:
    name = path.name.lower()
    if name in {"dockerfile", "makefile", "jenkinsfile"}:
        return True
    return path.suffix.lower() in SOURCE_SUFFIXES


def _source_file_priority(path: Path):
    suffix = path.suffix.lower()
    if suffix in PRIORITY_SOURCE_SUFFIXES:
        order = PRIORITY_SOURCE_SUFFIXES.index(suffix)
    else:
        order = len(PRIORITY_SOURCE_SUFFIXES) + 1
    return (order, len(path.parts), path.as_posix())


def _strip_line_comment_outside_quotes(line: str, marker: str) -> str:
    in_single = False
    in_double = False
    in_backtick = False
    escaped = False

    i = 0
    while i < len(line):
        ch = line[i]

        if escaped:
            escaped = False
            i += 1
            continue

        if ch == "\\":
            escaped = True
            i += 1
            continue

        if not in_double and not in_backtick and ch == "'":
            in_single = not in_single
            i += 1
            continue

        if not in_single and not in_backtick and ch == '"':
            in_double = not in_double
            i += 1
            continue

        if not in_single and not in_double and ch == "`":
            in_backtick = not in_backtick
            i += 1
            continue

        if (not in_single) and (not in_double) and (not in_backtick) and line.startswith(marker, i):
            return line[:i].rstrip()

        i += 1

    return line.rstrip()


def _strip_python_comments_and_docstrings(text: str) -> str:
    try:
        output = []
        prev_toktype = tokenize.INDENT
        last_lineno = 1
        last_col = 0

        for tok in tokenize.generate_tokens(io.StringIO(text).readline):
            toktype, tokstr, start, end, _ = tok
            start_line, start_col = start
            end_line, end_col = end

            if start_line > last_lineno:
                last_col = 0
            if start_col > last_col:
                output.append(" " * (start_col - last_col))

            is_docstring = (
                toktype == tokenize.STRING and
                prev_toktype in {tokenize.INDENT, tokenize.NEWLINE, tokenize.DEDENT}
            )

            if toktype == tokenize.COMMENT:
                pass
            elif is_docstring:
                pass
            else:
                output.append(tokstr)

            prev_toktype = toktype
            last_col = end_col
            last_lineno = end_line

        return "".join(output)
    except Exception:
        return text


def _strip_comments_from_text(path: Path, text: str) -> str:
    suffix = path.suffix.lower()
    name = path.name.lower()

    cleaned = text or ""

    if suffix == ".py":
        cleaned = _strip_python_comments_and_docstrings(cleaned)

    if suffix in {".html", ".xml", ".vue"}:
        cleaned = re.sub(r"<!--([\\s\\S]*?)-->", "", cleaned)

    if suffix in {".js", ".ts", ".tsx", ".jsx", ".java", ".kt", ".go", ".rs", ".cpp", ".cc", ".c", ".h", ".hpp", ".cs", ".php", ".swift", ".scala", ".css", ".scss", ".less", ".vue"}:
        cleaned = re.sub(r"/\*([\s\S]*?)\*/", "", cleaned)
        cleaned = "\n".join(_strip_line_comment_outside_quotes(line, "//") for line in cleaned.splitlines())

    if suffix in {".py", ".sh", ".bash", ".zsh", ".rb", ".yml", ".yaml", ".ini", ".toml", ".conf", ".env", ".dockerignore", ".gitignore"} or name in {"dockerfile", "makefile", "jenkinsfile"}:
        cleaned = "\n".join(_strip_line_comment_outside_quotes(line, "#") for line in cleaned.splitlines())

    if suffix == ".sql":
        cleaned = "\n".join(_strip_line_comment_outside_quotes(line, "--") for line in cleaned.splitlines())

    cleaned = "\n".join(line.rstrip() for line in cleaned.splitlines())
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _safe_extract_archive(archive_path: Path, output_dir: Path):


    ext = archive_path.suffix.lower()

    if ext == ".zip":
        with zipfile.ZipFile(archive_path, "r") as zf:
            zf.extractall(output_dir)
        return

    if ext in {".tar", ".gz", ".bz2", ".xz", ".tgz"} or archive_path.name.lower().endswith((".tar.gz", ".tar.bz2", ".tar.xz")):
        with tarfile.open(archive_path, "r:*") as tf:
            tf.extractall(output_dir)
        return

    if ext == ".7z":
        try:
            py7zr_module = importlib.import_module("py7zr")
        except Exception:
            raise HTTPException(status_code=400, detail="当前服务未安装 7z 解压依赖（py7zr），请改用 zip/tar 或安装 py7zr")

        with py7zr_module.SevenZipFile(archive_path, mode="r") as zf:
            zf.extractall(path=output_dir)
        return

    if ext == ".rar":
        try:
            rar_module = importlib.import_module("rarfile")
        except Exception:
            raise HTTPException(status_code=400, detail="当前服务未安装 rar 解压依赖（rarfile），请改用 zip/tar 或安装 rarfile")

        with rar_module.RarFile(archive_path) as rf:
            rf.extractall(path=output_dir)
        return


    try:
        shutil.unpack_archive(str(archive_path), str(output_dir))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"不支持的压缩格式或解压失败：{archive_path.name}，{str(e)}")


def _collect_source_text(extracted_dir: Path) -> str:
    chunks = []

    all_files = [p for p in extracted_dir.rglob("*") if p.is_file()]
    candidate_files = []
    for path in all_files:
        rel = path.relative_to(extracted_dir)
        rel_parts = set(rel.parts)
        if rel_parts & SKIP_DIR_NAMES:
            continue
        if not _is_supported_source_file(path):
            continue
        candidate_files.append(path)

    for path in sorted(candidate_files, key=_source_file_priority):
        rel = path.relative_to(extracted_dir)

        try:
            raw = path.read_bytes()
        except Exception:
            continue

        if b"\x00" in raw[:2048]:
            continue

        text = raw.decode("utf-8", errors="ignore")
        text = _strip_comments_from_text(path, text)
        if not text:
            continue

        item = f"\n### 文件: {rel.as_posix()}\n```\n{text}\n```\n"

        chunks.append(item)

    if not chunks:
        raise HTTPException(status_code=400, detail="压缩包内未识别到可评测源码文件")

    return "\n".join(chunks).strip()



def _extract_text_from_docx(report_path: Path) -> str:
    try:
        docx_module = importlib.import_module("docx")
        document = docx_module.Document(str(report_path))
        lines = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
        return "\n".join(lines).strip()
    except Exception:
        pass

    try:
        with zipfile.ZipFile(report_path, "r") as zf:
            xml_raw = zf.read("word/document.xml")
        root = ET.fromstring(xml_raw)
        lines: list[str] = []
        for p in root.iter():
            if not str(p.tag).endswith("}p"):
                continue
            segs = []
            for t in p.iter():
                if str(t.tag).endswith("}t") and t.text:
                    segs.append(t.text)
            line = "".join(segs).strip()
            if line:
                lines.append(line)
        return "\n".join(lines).strip()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"docx 文档解析失败：{str(e)}")



def _extract_text_from_pdf(report_path: Path) -> str:
    PdfReader = None
    try:
        PdfReader = importlib.import_module("pypdf").PdfReader
    except Exception:
        try:
            PdfReader = importlib.import_module("PyPDF2").PdfReader
        except Exception:
            pass

    if not PdfReader:
        raise HTTPException(status_code=400, detail="当前服务未安装 PDF 文本提取依赖（pypdf/PyPDF2）")

    try:
        reader = PdfReader(str(report_path))
        pages = []
        for page in reader.pages:
            text = page.extract_text() if hasattr(page, "extract_text") else ""
            text = (text or "").strip()
            if text:
                pages.append(text)
        return "\n\n".join(pages).strip()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"pdf 文档解析失败：{str(e)}")



def _extract_text_from_doc(report_path: Path) -> str:
    try:
        textract_module = importlib.import_module("textract")
        raw = textract_module.process(str(report_path))
        return (raw or b"").decode("utf-8", errors="ignore").strip()
    except Exception:
        raise HTTPException(status_code=400, detail="当前服务暂不支持 .doc 直接抽取，请转为 docx/pdf 后重传")



def _resolve_report_document_content(file_path: str) -> str:
    path_text = (file_path or "").strip()
    if not path_text:
        raise HTTPException(status_code=400, detail="报告文件路径为空")

    report_path = Path(path_text)
    if (not report_path.exists()) or (not report_path.is_file()):
        raise HTTPException(status_code=400, detail="报告文件不存在，请让学生重新提交")

    suffix = report_path.suffix.lower()
    if suffix == ".docx":
        content = _extract_text_from_docx(report_path)
    elif suffix == ".pdf":
        content = _extract_text_from_pdf(report_path)
    elif suffix == ".doc":
        content = _extract_text_from_doc(report_path)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的报告文件类型：{suffix}")

    content = (content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="报告文件未解析到有效正文")

    if len(content) > 120000:
        content = _shrink_text_keep_head_tail(content, 120000)

    return f"### 报告正文（自动抽取）\n{content}".strip()



def _resolve_submission_code_content(code_content: str) -> str:

    text = (code_content or "").strip()
    if not text:
        return ""

    if text.startswith(REPORT_FILE_MARKER):
        report_file_path = text[len(REPORT_FILE_MARKER):].strip()
        return _resolve_report_document_content(report_file_path)

    if not text.startswith(ARCHIVE_MARKER):
        return text

    archive_path_str = text[len(ARCHIVE_MARKER):].strip()
    archive_path = Path(archive_path_str)
    if (not archive_path.exists()) or (not archive_path.is_file()):
        raise HTTPException(status_code=400, detail="提交的压缩包文件不存在，请让学生重新提交")

    try:
        with tempfile.TemporaryDirectory(prefix="ai_review_extract_") as tmp_dir:
            extract_dir = Path(tmp_dir)
            _safe_extract_archive(archive_path, extract_dir)
            return _collect_source_text(extract_dir)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"压缩包处理失败：{str(e)}")



def _report_ai_history_root_dir() -> Path:
    root = Path(__file__).resolve().parent / "document" / REPORT_AI_HISTORY_DIRNAME
    root.mkdir(parents=True, exist_ok=True)
    return root



def _safe_history_filename_part(text: str, fallback: str) -> str:
    raw = (text or "").strip()
    cleaned = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", raw)
    cleaned = cleaned.strip("_")
    return cleaned or fallback



def _build_report_student_history_path(db: Session, report_submission: models.ReportSubmission) -> Path:
    student = db.query(models.User).filter(models.User.id == report_submission.student_id).first()
    student_no_raw = getattr(student, "student_no", None)
    student_real_name = getattr(student, "real_name", None)
    student_username = getattr(student, "username", None)

    student_no = _safe_history_filename_part(str(student_no_raw or ""), f"sid_{report_submission.student_id}")
    student_name = _safe_history_filename_part(str(student_real_name or student_username or ""), f"student_{report_submission.student_id}")
    return _report_ai_history_root_dir() / f"{student_no}_{student_name}_{report_submission.student_id}.md"




def _append_report_ai_history_record(
    db: Session,
    report_submission: models.ReportSubmission,
    ai_model: str,
    ai_feedback: str,
    score: int,
    checks: str,
    history_id: int,
):
    md_path = _build_report_student_history_path(db, report_submission)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    model_text = (ai_model or "").strip() or "未知模型"
    checks_text = (checks or "").strip() or "（未设置）"
    feedback_text = (ai_feedback or "").strip()

    block = (
        f"\n<!-- {REPORT_AI_RECORD_START}:{history_id} -->\n"
        f"## AI评测记录 {now}\n"
        f"- 记录ID：{history_id}\n"
        f"- 报告提交ID：{report_submission.id}\n"
        f"- 学生ID：{report_submission.student_id}\n"
        f"- 模型：{model_text}\n"
        f"- 评分：{score}\n"
        f"- 时间：{now}\n\n"
        f"### 批改要点\n"
        f"{checks_text}\n\n"
        f"### AI评语\n"
        f"{feedback_text}\n"
        f"<!-- {REPORT_AI_RECORD_END}:{history_id} -->\n"
    )

    if md_path.exists():
        original = md_path.read_text(encoding="utf-8", errors="ignore")
        content = (original.rstrip() + "\n\n" + block.lstrip()) if original.strip() else block.lstrip()
    else:
        header = "# 报告AI批改历史\n\n"
        content = header + block.lstrip()

    md_path.write_text(content, encoding="utf-8")



def _parse_report_ai_history_records(md_path: Path):
    if not md_path.exists():
        return []

    text = md_path.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        rf"<!--\s*{REPORT_AI_RECORD_START}:(\d+)\s*-->([\s\S]*?)<!--\s*{REPORT_AI_RECORD_END}:\1\s*-->",
        flags=re.S,
    )

    rows = []
    for m in pattern.finditer(text):
        record_id = int(m.group(1))
        body = (m.group(2) or "").strip()

        model_m = re.search(r"-\s*模型\s*[：:]\s*(.+)", body)
        score_m = re.search(r"-\s*评分\s*[：:]\s*(\d{1,3})", body)
        time_m = re.search(r"-\s*时间\s*[：:]\s*(.+)", body)
        feedback_m = re.search(r"###\s*AI评语\s*\n([\s\S]*)$", body, flags=re.S)

        rows.append({
            "record_id": record_id,
            "ai_model": model_m.group(1).strip() if model_m else None,
            "score": int(score_m.group(1)) if score_m else None,
            "created_at": time_m.group(1).strip() if time_m else None,
            "review_text": feedback_m.group(1).strip() if feedback_m else body,
        })

    rows.sort(key=lambda x: x["record_id"], reverse=True)
    return rows



def _delete_report_ai_history_record(md_path: Path, record_id: int) -> bool:
    if not md_path.exists():
        return False

    text = md_path.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        rf"\n?<!--\s*{REPORT_AI_RECORD_START}:{record_id}\s*-->[\s\S]*?<!--\s*{REPORT_AI_RECORD_END}:{record_id}\s*-->\n?",
        flags=re.S,
    )

    new_text, count = pattern.subn("\n", text, count=1)
    if count <= 0:
        return False

    new_text = re.sub(r"\n{3,}", "\n\n", new_text).strip() + "\n"
    md_path.write_text(new_text, encoding="utf-8")
    return True



def _clip_code_for_prompt(code_text: str) -> str:

    return (code_text or "").strip()




def _is_retryable_model_error(err: Exception) -> bool:
    if isinstance(err, IncompleteRead):
        return True
    if isinstance(err, urllib.error.URLError):
        reason = str(err.reason).lower()
        return any(x in reason for x in ["timed out", "connection reset", "temporarily", "remote end closed"])
    msg = str(err).lower()
    return any(x in msg for x in ["incompleteread", "timed out", "connection reset", "remote end closed"])


def _build_model_request(url: str, api_key: str, payload: dict[str, Any]) -> urllib.request.Request:
    return urllib.request.Request(
        url=url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "Accept": "application/json",
            "Accept-Encoding": "identity",
            "Connection": "close"
        },
        method="POST"
    )


def _shrink_prompt_content_for_retry(content: str, attempt: int) -> str:
    text = content or ""
    if not text or attempt <= 0:
        return text

    target_len = max(
        MIN_RETRY_PROMPT_CHARS,
        int(len(text) * (RETRY_PROMPT_SHRINK_RATIO ** attempt))
    )
    if len(text) <= target_len:
        return text

    head_len = int(target_len * 0.82)
    tail_len = max(0, target_len - head_len)
    head = text[:head_len]
    tail = text[-tail_len:] if tail_len > 0 else ""
    return f"{head}\n\n...（重试降载：中间部分源码已省略）...\n\n{tail}"


def _resolve_model_limits(model_name: str) -> tuple[int, int, int]:
    name = (model_name or "").strip().lower()

    if "deepseek" in name:
        if "reasoner" in name or "r1" in name:
            return DEEPSEEK_CONTEXT_WINDOW_TOKENS, DEEPSEEK_REASONER_MAX_OUTPUT_TOKENS, CONTEXT_WINDOW_RESERVED_TOKENS
        return DEEPSEEK_CONTEXT_WINDOW_TOKENS, DEEPSEEK_CHAT_MAX_OUTPUT_TOKENS, CONTEXT_WINDOW_RESERVED_TOKENS

    if "qwen" in name:
        # Qwen 模型通常对 max_tokens 限制较严 (最大通常为 8192 或 4096)
        return ASSUMED_CONTEXT_WINDOW_TOKENS, 8192, CONTEXT_WINDOW_RESERVED_TOKENS

    if "doubao" in name:
        # Doubao 模型对 max_tokens 的限制通常为 4096
        return ASSUMED_CONTEXT_WINDOW_TOKENS, 4096, CONTEXT_WINDOW_RESERVED_TOKENS

    return ASSUMED_CONTEXT_WINDOW_TOKENS, MAX_BACKEND_TOKENS, CONTEXT_WINDOW_RESERVED_TOKENS


def _prompt_soft_limit_tokens(model_name: str) -> int:
    name = (model_name or "").strip().lower()
    if "deepseek" in name:
        if "reasoner" in name or "r1" in name:
            return DEEPSEEK_REASONER_PROMPT_SOFT_LIMIT_TOKENS
        return DEEPSEEK_CHAT_PROMPT_SOFT_LIMIT_TOKENS
    return DEFAULT_PROMPT_SOFT_LIMIT_TOKENS


def _recommended_prompt_limit(model_name: str) -> int:
    context_window, output_cap, reserved = _resolve_model_limits(model_name)
    target_output = min(8192, output_cap)
    theoretical_limit = max(12000, context_window - reserved - target_output)
    return min(theoretical_limit, _prompt_soft_limit_tokens(model_name))


def _shrink_text_keep_head_tail(text: str, target_chars: int) -> str:
    src = text or ""
    limit = max(MIN_CLIPPED_CODE_CHARS, int(target_chars or 0))
    if len(src) <= limit:
        return src

    head_len = int(limit * 0.8)
    tail_len = max(0, limit - head_len)
    head = src[:head_len]
    tail = src[-tail_len:] if tail_len > 0 else ""
    return f"{head}\n\n...（自动降载：中间源码已省略）...\n\n{tail}"


def _adaptive_completion_max_tokens(requested: int, prompt_tokens: int, model_name: str) -> int:
    requested_max = min(MAX_BACKEND_TOKENS, max(256, int(requested or 100000)))
    context_window, model_output_cap, reserved = _resolve_model_limits(model_name)

    by_context_budget = max(
        MIN_RETRY_MAX_TOKENS,
        context_window - max(0, int(prompt_tokens)) - reserved
    )
    safe_cap = min(requested_max, model_output_cap, by_context_budget)

    soft_limit = _prompt_soft_limit_tokens(model_name)
    if prompt_tokens >= soft_limit:
        safe_cap = min(safe_cap, 1024)
    elif prompt_tokens >= int(soft_limit * 0.85):
        safe_cap = min(safe_cap, 1280)
    elif prompt_tokens >= int(soft_limit * 0.7):
        safe_cap = min(safe_cap, 1536)

    usage_ratio = max(0.0, min(1.0, float(prompt_tokens) / float(max(1, context_window))))
    if usage_ratio >= 0.85:
        safe_cap = min(safe_cap, 1024)
    elif usage_ratio >= 0.75:
        safe_cap = min(safe_cap, 1536)

    return max(256, int(safe_cap))





def _request_model_with_retry(url: str, api_key: str, payload: dict[str, Any]):
    base_max_tokens = min(
        MAX_BACKEND_TOKENS,
        max(256, int(payload.get("max_tokens") or 1024))
    )

    for attempt in range(MAX_MODEL_RETRIES + 1):
        raw = ""
        attempt_payload = dict(payload)

        attempt_max_tokens = max(
            MIN_RETRY_MAX_TOKENS,
            int(base_max_tokens * (RETRY_MAX_TOKENS_RATIO ** attempt))
        )
        attempt_payload["max_tokens"] = min(base_max_tokens, attempt_max_tokens)

        msgs = list(attempt_payload.get("messages") or [])
        if attempt > 0 and len(msgs) >= 2 and isinstance(msgs[1], dict):
            msgs[1] = dict(msgs[1])
            msgs[1]["content"] = _shrink_prompt_content_for_retry(str(msgs[1].get("content") or ""), attempt)
            attempt_payload["messages"] = msgs

        attempt_payload["stream"] = True
        if "stream_options" not in attempt_payload:
            attempt_payload["stream_options"] = {"include_usage": True}
            
        full_content = []
        full_reasoning = []

        try:
            request = _build_model_request(url, api_key, attempt_payload)
            with urllib.request.urlopen(request, timeout=300) as resp:
                content_type = resp.headers.get("Content-Type", "")
                if "application/json" in content_type.lower():
                    raw = resp.read().decode("utf-8", errors="ignore").strip()
                    if not raw:
                        raise ValueError("empty response")
                    return json.loads(raw)

                final_usage = None
                
                for line in resp:
                    line_str = line.decode("utf-8", errors="ignore").strip()
                    if not line_str or line_str.startswith(":"):
                        continue
                    if line_str == "data: [DONE]":
                        break
                    if line_str.startswith("data: "):
                        try:
                            chunk = json.loads(line_str[6:])
                            choices = chunk.get("choices", [])
                            if choices and isinstance(choices, list):
                                delta = choices[0].get("delta", {})
                                if "content" in delta and delta["content"]:
                                    full_content.append(delta["content"])
                                if "reasoning_content" in delta and delta["reasoning_content"]:
                                    full_reasoning.append(delta["reasoning_content"])
                            
                            usage = chunk.get("usage")
                            if usage and isinstance(usage, dict):
                                final_usage = usage
                        except json.JSONDecodeError:
                            pass

                content_str = "".join(full_content)
                if not content_str and not full_reasoning:
                    raise ValueError("empty response")
                
                msg = {"content": content_str}
                if full_reasoning:
                    msg["reasoning_content"] = "".join(full_reasoning)
                
                result = {"choices": [{"message": msg}]}
                if final_usage:
                    result["usage"] = final_usage
                
                return result

        except IncompleteRead as e:
            # 如果流式读取中断，尝试返回已累积的部分内容
            partial_content = "".join(full_content)
            if partial_content:
                msg = {"content": partial_content}
                if full_reasoning:
                    msg["reasoning_content"] = "".join(full_reasoning)
                return {"choices": [{"message": msg}]}
            
            if attempt >= MAX_MODEL_RETRIES:
                raise HTTPException(status_code=502, detail="模型响应中断（IncompleteRead），已多次重试仍失败，请缩小代码体积后重试")
            time.sleep(MODEL_RETRY_BACKOFF_SECONDS * (2 ** attempt))

        except json.JSONDecodeError:
            recovered = _recover_text_from_truncated_json(raw if "raw" in locals() else "")
            if recovered:
                return {"choices": [{"message": {"content": recovered}}]}
            if attempt >= MAX_MODEL_RETRIES:
                raise HTTPException(status_code=502, detail="模型返回内容被截断，已多次重试仍失败，请缩小代码体积后重试")
            time.sleep(MODEL_RETRY_BACKOFF_SECONDS * (2 ** attempt))


        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else ""
            detail = err_body.strip() or str(e)
            if e.code in {400, 408, 409, 425, 429, 500, 502, 503, 504} and attempt < MAX_MODEL_RETRIES:
                time.sleep(MODEL_RETRY_BACKOFF_SECONDS * (2 ** attempt))
                continue
            raise HTTPException(status_code=400, detail=f"模型调用失败：{detail}")

        except urllib.error.URLError as e:
            if _is_retryable_model_error(e) and attempt < MAX_MODEL_RETRIES:
                time.sleep(MODEL_RETRY_BACKOFF_SECONDS * (2 ** attempt))
                continue
            raise HTTPException(status_code=400, detail=f"模型调用失败：{str(e.reason)}")

        except Exception as e:
            if _is_retryable_model_error(e) and attempt < MAX_MODEL_RETRIES:
                time.sleep(MODEL_RETRY_BACKOFF_SECONDS * (2 ** attempt))
                continue
            if "IncompleteRead" in str(e):
                raise HTTPException(status_code=502, detail="模型响应中断（IncompleteRead），请重试或缩小代码体积")
            raise HTTPException(status_code=500, detail=f"模型调用异常：{str(e)}")

    raise HTTPException(status_code=502, detail="模型响应异常，请稍后重试")




@app.post("/ai/review")
def run_real_ai_review(req: AIReviewRequest, db: Session = Depends(get_db)):




    provider = req.provider
    base_url = (provider.base_url or "").strip()
    api_key = (provider.api_key or "").strip()
    model_name = (provider.model_name or "").strip()

    if (not base_url) or (not api_key) or (not model_name):
        raise HTTPException(status_code=400, detail="请填写该模型API数据")

    checks = (req.functionality_checks or "").strip()
    if not checks:
        checks = "1. 核心功能正确性\n2. 异常处理与边界情况\n3. 代码规范与可维护性"

    task_type = (req.task_type or "assignment").strip().lower()
    assignment_prompt = (req.assignment_prompt_template or "").strip()
    report_prompt = (req.report_prompt_template or "").strip()

    prompt_template = report_prompt if task_type == "report" else assignment_prompt
    if not prompt_template:
        prompt_template = (req.prompt_template or "").strip()

    if not prompt_template:
        if task_type == "report":
            prompt_template = "请根据{functionality_checks}对学生报告进行结构化评分与反馈，并给出总分(0-100)和改进建议。"
        else:
            prompt_template = "请根据{functionality_checks}对学生代码进行评分与反馈，并严格输出三段评分格式。"

    if "{functionality_checks}" in prompt_template:
        final_prompt = prompt_template.replace("{functionality_checks}", checks)
    else:
        final_prompt = f"{prompt_template}\n\n【本次重点检查项】\n{checks}"

    if len(final_prompt) > MAX_PROMPT_INSTRUCTION_CHARS:
        final_prompt = final_prompt[:MAX_PROMPT_INSTRUCTION_CHARS] + "\n\n...（评测要求过长，已截断）"

    resolved_code_content = _resolve_submission_code_content(req.code_content)
    prompt_code_content = _clip_code_for_prompt(resolved_code_content)

    safe_temperature = max(0.01, min(1.0, float(req.temperature)))

    def _build_messages_with_code(code_content: str) -> list[dict[str, str]]:
        user_prompt_text = f"{final_prompt}\n\n【学生提交内容】\n{code_content}"
        return [
            {"role": "system", "content": "你是编程作业评测助手，请按给定格式完整但精炼地输出评测结果，避免冗长复述代码。"},
            {"role": "user", "content": user_prompt_text}
        ]

    messages = _build_messages_with_code(prompt_code_content)
    prompt_tokens_estimated = _estimate_messages_tokens(messages)

    soft_prompt_limit = _prompt_soft_limit_tokens(model_name)
    trim_rounds = 0
    while prompt_tokens_estimated > soft_prompt_limit and prompt_code_content and trim_rounds < 3:
        code_tokens = max(1, _estimate_text_tokens(prompt_code_content))
        overflow_tokens = prompt_tokens_estimated - soft_prompt_limit
        target_code_tokens = max(1200, code_tokens - overflow_tokens - 512)
        shrink_ratio = max(0.2, min(0.95, float(target_code_tokens) / float(code_tokens)))
        target_chars = int(len(prompt_code_content) * shrink_ratio)
        prompt_code_content = _shrink_text_keep_head_tail(prompt_code_content, target_chars)
        messages = _build_messages_with_code(prompt_code_content)
        prompt_tokens_estimated = _estimate_messages_tokens(messages)
        trim_rounds += 1

    safe_max_tokens = _adaptive_completion_max_tokens(int(req.max_tokens or 100000), prompt_tokens_estimated, model_name)



    url = _build_chat_completions_url(base_url)
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": safe_temperature,
        "max_tokens": safe_max_tokens,
        "stream": False
    }





    try:
        data = _request_model_with_retry(url, api_key, payload)
    except HTTPException as e:
        context_window, _, _ = _resolve_model_limits(model_name)
        recommended_prompt_limit = _recommended_prompt_limit(model_name)
        raise HTTPException(
            status_code=e.status_code,
            detail=f"{str(e.detail)}（本次提交Prompt约 {prompt_tokens_estimated} tokens；当前模型上下文总长约 {context_window} tokens，建议Prompt控制在 {recommended_prompt_limit} tokens 以内）"
        )





    ai_feedback = _extract_ai_text(data)



    if not ai_feedback:
        raise HTTPException(status_code=400, detail="模型返回为空，请检查模型配置或接口权限")

    if req.submission_id:
        history = models.AIReviewHistory(
            submission_id=req.submission_id,
            assignment_id=req.assignment_id,
            ai_model=req.ai_model,
            review_text=ai_feedback
        )
        db.add(history)
        db.commit()
        db.refresh(history)

        if task_type == "report":
            report_submission = db.query(models.ReportSubmission).filter(
                models.ReportSubmission.id == req.submission_id
            ).first()
            history_id = int(getattr(history, "id", 0) or 0)
            if report_submission and history_id > 0:
                _append_report_ai_history_record(
                    db=db,
                    report_submission=report_submission,
                    ai_model=req.ai_model,
                    ai_feedback=ai_feedback,
                    score=_parse_total_score(ai_feedback),
                    checks=checks,
                    history_id=history_id,
                )



    usage = data.get("usage") if isinstance(data, dict) else None
    provider_prompt_tokens = usage.get("prompt_tokens") if isinstance(usage, dict) else None
    prompt_tokens_submitted = int(provider_prompt_tokens) if isinstance(provider_prompt_tokens, (int, float)) else prompt_tokens_estimated

    return {
        "message": f"已成功调用 {req.ai_model} 模型",
        "task_type": task_type,
        "ai_feedback": ai_feedback,
        "score": _parse_total_score(ai_feedback),
        "final_prompt": final_prompt,
        "prompt_tokens_submitted": prompt_tokens_submitted,
        "prompt_tokens_source": "provider" if isinstance(provider_prompt_tokens, (int, float)) else "estimated",
        "max_tokens_submitted": safe_max_tokens
    }





@app.get("/ai/review/history/{submission_id}")
def get_ai_review_history(
    submission_id: int,
    limit: int = Query(10, ge=1, le=100),
    ai_model: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(models.AIReviewHistory).filter(
        models.AIReviewHistory.submission_id == submission_id
    )
    if ai_model:
        query = query.filter(models.AIReviewHistory.ai_model == ai_model)

    rows = query.order_by(models.AIReviewHistory.id.desc()).limit(limit).all()

    return [
        {
            "id": x.id,
            "submission_id": x.submission_id,
            "assignment_id": x.assignment_id,
            "ai_model": x.ai_model,
            "review_text": x.review_text,
            "created_at": x.created_at.isoformat() if x.created_at else None
        }
        for x in rows
    ]


@app.delete("/ai/review/history/{submission_id}")
def clear_ai_review_history(submission_id: int, db: Session = Depends(get_db)):
    deleted = db.query(models.AIReviewHistory).filter(
        models.AIReviewHistory.submission_id == submission_id
    ).delete(synchronize_session=False)
    db.commit()
    return {"message": "已清空当前提交的AI评测历史", "deleted_count": deleted}



# ================= ⚡ 新增：老师正式发布学情报告 (存入数据库) ⚡ =================

# 🌟 修改：批改接口现在接收老师临时输入的 AI 评测要点
@app.post("/publish_report/")
def publish_report(data: schemas.ReportCreate, db: Session = Depends(get_db)):
    sub = db.query(models.Submission).filter(models.Submission.id == data.submission_id).first()
    if not sub: raise HTTPException(status_code=404, detail="未找到提交记录")
    
    sub.ai_feedback = data.ai_feedback
    sub.teacher_comment = data.teacher_comment
    sub.score = data.score
    # 🌟 建议将老师本次输入的要点也存下来，方便学生看报告时知道 AI 是按什么标准评的
    sub.ai_criteria = data.ai_criteria 
    sub.status = "finished"
    db.commit()
    return {"message": "批改结果已发布"}

# ================= ⚡ 新增：学生实时查询自己的学情报告 ⚡ =================
# 🌟 修改：根据学生 ID 和 作业 ID 精确查询报告
@app.get("/my_report/{student_id}/{assignment_id}")
def get_my_report(student_id: int, assignment_id: int, db: Session = Depends(get_db)):
    # 增加了一个过滤条件：assignment_id
    sub = db.query(models.Submission).filter(
        models.Submission.student_id == student_id,
        models.Submission.assignment_id == assignment_id
    ).order_by(models.Submission.id.desc()).first()
    
    if not sub:
        return {"status": "none"}
    
    return {
        "status": sub.status,
        "ai_feedback": sub.ai_feedback,
        "teacher_comment": sub.teacher_comment,
        "score": sub.score
    }

# 🌟 获取所有作业列表
@app.get("/assignments/all")
def get_all_assignments(db: Session = Depends(get_db)):
    return db.query(models.Assignment).order_by(models.Assignment.id.desc()).all()

# 🌟 根据 ID 获取特定作业详情
@app.get("/assignments/{assignment_id}")
def get_assignment_detail(assignment_id: int, db: Session = Depends(get_db)):
    assign = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not assign:
        raise HTTPException(status_code=404, detail="作业不存在")
    return assign

# 🌟 获取特定作业的所有学生提交记录
@app.get("/assignments/{assignment_id}/submissions")
def get_assignment_submissions(assignment_id: int, db: Session = Depends(get_db)):
    assignment = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="作业不存在")

    class_id = assignment.course_id
    target_type = _get_assignment_target_type(db, assignment_id, class_id=class_id)
    subs = db.query(models.Submission).filter(
        models.Submission.assignment_id == assignment_id
    ).order_by(models.Submission.id.desc()).all()

    if target_type == "group" and class_id:
        teams = db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).all()
        team_map = {t.id: t for t in teams}

        member_rows = db.query(models.ClassTeamMember).filter(
            models.ClassTeamMember.team_id.in_(list(team_map.keys()))
        ).all() if team_map else []
        student_ids = [x.student_id for x in member_rows]
        users = db.query(models.User).filter(models.User.id.in_(student_ids)).all() if student_ids else []
        user_map = {u.id: u for u in users}

        team_members: dict[int, list[str]] = {tid: [] for tid in team_map.keys()}
        for m in member_rows:
            u = user_map.get(m.student_id)
            if not u:
                continue
            team_members[m.team_id].append(u.real_name or u.username)

        valid_team_ids = [tid for tid, names in team_members.items() if names]
        meta_rows = db.query(models.SubmissionTeamMeta).filter(
            models.SubmissionTeamMeta.submission_id.in_([s.id for s in subs])
        ).all() if subs else []
        submission_team_map = {m.submission_id: m.team_id for m in meta_rows if m.team_id}

        latest_team_sub: dict[int, models.Submission] = {}
        for s in subs:
            team_id = submission_team_map.get(s.id)
            if team_id in valid_team_ids and (team_id not in latest_team_sub):
                latest_team_sub[team_id] = s

        result = []
        for team_id in valid_team_ids:
            team = team_map.get(team_id)
            if not team:
                continue

            latest = latest_team_sub.get(team_id)
            if not latest:
                status = "unsubmitted"
            else:
                status = latest.status if latest.status in ("submitted", "finished") else "submitted"

            result.append({
                "id": latest.id if latest else -team_id,
                "student_id": -team_id,
                "student_name": f"{team.name}（小组）",
                "student_no": f"{len(team_members.get(team_id, []))}人",
                "team_id": team_id,
                "team_name": team.name,
                "team_members": team_members.get(team_id, []),
                "status": status,
                "score": latest.score if latest and latest.score is not None else None,
                "code_content": latest.code_content if latest else None,
                "ai_feedback": latest.ai_feedback if latest else None,
                "teacher_comment": latest.teacher_comment if latest else None,
                "ai_criteria": (latest.ai_criteria if latest and latest.ai_criteria else assignment.ai_criteria),
                "submission_id": latest.id if latest else None,
                "target_type": "group",
                "created_at": "2026-03-11"
            })
        return result

    if class_id:
        target_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
        students = _get_students_for_class(db, target_class) if target_class else []
        student_ids = [s.id for s in students]

        latest_student_sub: dict[int, models.Submission] = {}
        for s in subs:
            if s.student_id in student_ids and s.student_id not in latest_student_sub:
                latest_student_sub[s.student_id] = s

        result = []
        for stu in students:
            latest = latest_student_sub.get(stu.id)
            if not latest:
                status = "unsubmitted"
            else:
                status = latest.status if latest.status in ("submitted", "finished") else "submitted"

            result.append({
                "id": latest.id if latest else -stu.id,
                "student_id": stu.id,
                "student_name": stu.real_name or stu.username,
                "student_no": stu.student_no,
                "status": status,
                "score": latest.score if latest and latest.score is not None else None,
                "code_content": latest.code_content if latest else None,
                "ai_feedback": latest.ai_feedback if latest else None,
                "teacher_comment": latest.teacher_comment if latest else None,
                "ai_criteria": (latest.ai_criteria if latest and latest.ai_criteria else assignment.ai_criteria),
                "submission_id": latest.id if latest else None,
                "target_type": "individual",
                "created_at": "2026-03-11"
            })
        return result

    result = []
    for s in subs:
        user = db.query(models.User).filter(models.User.id == s.student_id).first()
        result.append({
            "id": s.id,
            "student_id": s.student_id,
            "student_name": (user.real_name or user.username) if user else "未知学生",
            "student_no": user.student_no if user else None,
            "status": s.status,
            "score": s.score,
            "code_content": s.code_content,
            "ai_feedback": s.ai_feedback,
            "teacher_comment": s.teacher_comment,
            "ai_criteria": s.ai_criteria,
            "submission_id": s.id,
            "target_type": "individual",
            "created_at": "2026-03-11"
        })
    return result


# 🌟 老师/管理员专用：获取学生平均分排名
@app.get("/analytics/student_ranking")
def get_student_ranking(db: Session = Depends(get_db)):
    # 查询已完成批改的提交记录，按学生 ID 分组，计算平均分
    rankings = db.query(
        models.User.username,
        models.User.real_name,
        func.avg(models.Submission.score).label('avg_score'),
        func.count(models.Submission.id).label('submit_count')
    ).join(models.Submission, models.User.id == models.Submission.student_id)\
     .filter(models.Submission.status == "finished")\
     .group_by(models.User.username, models.User.real_name)\
     .order_by(func.avg(models.Submission.score).desc())\
     .all()

    return [
        {"name": r.real_name or r.username, "score": round(float(r.avg_score), 2), "count": r.submit_count} 
        for r in rankings
    ]

# --- main.py ---

# 🌟 更新个人资料接口
@app.post("/users/profile/update")
def update_profile(data: schemas.ProfileUpdate, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == data.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    
    user.bio = data.bio
    user.skills = data.skills
    db.commit()
    return {"message": "个人资料更新成功！"}

# 🌟 获取组队推荐接口 (简单互补算法)
@app.get("/teams/recommendations/{user_id}")
def get_recommendations(user_id: int, db: Session = Depends(get_db)):
    me = db.query(models.User).filter(models.User.id == user_id).first()
    # 查找所有学生
    all_students = db.query(models.User).filter(
        models.User.role == "student", 
        models.User.id != user_id
    ).all()
    
    recommends = []
    my_skills = (me.skills or "").lower()

    for s in all_students:
        s_skills = (s.skills or "").lower()
        # 简单互补逻辑：如果我擅长后端(python)，他擅长前端(vue)，则匹配度高
        match_rate = 60
        reason = "背景相似"
        
        if "python" in my_skills and "vue" in s_skills:
            match_rate = 95
            reason = "技术栈完美互补 (后端 + 前端)"
        elif "algorithm" in my_skills and "develop" in s_skills:
            match_rate = 90
            reason = "能力互补 (算法 + 开发)"
            
        recommends.append({
            "id": s.id,
            "name": s.real_name or s.username,
            "skills": s.skills or "暂未填写",
            "match_rate": f"{match_rate}%",
            "reason": reason
        })
    
    # 按匹配度排序返回前3名
    return sorted(recommends, key=lambda x: x['match_rate'], reverse=True)[:3]

@app.get("/analytics/student_profile/{user_id}")
def get_real_profile(user_id: int, db: Session = Depends(get_db)):
    # 1. 获取该学生所有已批改的提交记录
    submissions = db.query(models.Submission).filter(
        models.Submission.student_id == user_id,
        models.Submission.status == "finished"
    ).all()

    # 默认值（如果没有数据）
    default_stats = [60, 60, 60, 60, 60]

    if not submissions:
        return {"radar_data": default_stats, "message": "暂无作业数据"}

    # 2. 计算各项指标（这里是你的“算法”逻辑）
    # 平均分作为核心基准
    avg_score = sum([s.score for s in submissions]) / len(submissions)
    
    # 维度1：算法逻辑 (由平均分决定)
    logic_score = avg_score
    
    # 维度2：代码规范 (模拟逻辑：如果 ai_feedback 提到“规范”，则加分)
    # 实际应用中，你可以根据数据库里的 ai_criteria 字段来判定
    clean_code_score = min(100, 60 + (len(submissions) * 5)) 
    
    # 维度3：工程能力 (由提交次数决定，提交越多代表实践越多)
    eng_score = min(100, 50 + (len(submissions) * 10))
    
    # 维度4：按时提交 (这里可以对比 deadline，暂给固定高分)
    on_time_score = 95
    
    # 维度5：AI 交互活跃度 (模拟逻辑)
    ai_active_score = 85

    return {
        "radar_data": [
            round(clean_code_score, 1), # 代码规范
            round(logic_score, 1),      # 算法逻辑
            round(eng_score, 1),        # 工程能力
            round(on_time_score, 1),    # 按时提交
            round(ai_active_score, 1)   # AI 交互
        ]
    }

# ================= 学期/班级/学生/队伍持久化接口 =================
def _get_current_semester(db: Session):
    return db.query(models.Semester).filter(models.Semester.is_current == True).first()


def _normalize_class_name(class_name: Optional[str]) -> Optional[str]:
    text = (class_name or "").strip()
    return text or None



def _ensure_student_class_relation(
    db: Session,
    student: models.User,
    semester_id: Optional[int] = None,
):
    if not student or student.role != "student":
        return None, False

    changed = False
    class_name = _normalize_class_name(student.class_name)
    if class_name and student.class_name != class_name:
        student.class_name = class_name
        changed = True

    canonical_teacher_id = _canonical_teacher_id(db, student.teacher_id)
    if canonical_teacher_id is not None and student.teacher_id != canonical_teacher_id:
        student.teacher_id = canonical_teacher_id
        changed = True

    if canonical_teacher_id is None or not class_name:
        return None, changed

    if semester_id is None:
        current = _get_current_semester(db)
        if not current:
            return None, changed
        semester_id = current.id

    teaching_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.semester_id == semester_id,
        models.TeachingClass.teacher_id == canonical_teacher_id,
        models.TeachingClass.name == class_name
    ).first()
    if not teaching_class:
        teaching_class = models.TeachingClass(
            name=class_name,
            semester_id=semester_id,
            teacher_id=canonical_teacher_id,
        )
        db.add(teaching_class)
        db.flush()
        changed = True

    relation = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == teaching_class.id,
        models.ClassStudent.student_id == student.id
    ).first()
    if not relation:
        db.add(models.ClassStudent(class_id=teaching_class.id, student_id=student.id))
        changed = True

    return teaching_class, changed



def _sync_teacher_classes_from_students(db: Session, semester_id: int, teacher_id: Optional[int] = None):
    q = db.query(models.User).filter(
        models.User.role == "student",
        models.User.teacher_id.isnot(None),
        models.User.class_name.isnot(None),
        func.length(func.trim(models.User.class_name)) > 0
    )
    if teacher_id is not None:
        canonical_teacher_id = _canonical_teacher_id(db, teacher_id)
        if canonical_teacher_id is None:
            return
        q = q.filter(models.User.teacher_id == canonical_teacher_id)

    students = q.order_by(models.User.id.asc()).all()
    changed = False
    for stu in students:
        _, student_changed = _ensure_student_class_relation(db, stu, semester_id=semester_id)
        changed = changed or student_changed

    if changed:
        db.commit()



def _repair_current_semester_class_data():
    db = SessionLocal()
    try:
        current = _get_current_semester(db)
        if current:
            _sync_teacher_classes_from_students(db, current.id)
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()



_repair_current_semester_class_data()



@app.get("/semesters")
def list_semesters(db: Session = Depends(get_db)):
    semesters = db.query(models.Semester).order_by(models.Semester.id.desc()).all()
    return [
        {
            "id": s.id,
            "name": s.name,
            "start_date": s.start_date,
            "end_date": s.end_date,
            "is_enabled": s.is_enabled,
            "is_current": s.is_current,
        }
        for s in semesters
    ]


@app.get("/semesters/current")
def get_current_semester(db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        return {"id": None, "name": None, "start_date": None, "end_date": None, "is_enabled": False}
    return {
        "id": current.id,
        "name": current.name,
        "start_date": current.start_date,
        "end_date": current.end_date,
        "is_enabled": current.is_enabled,
    }


@app.post("/admin/semesters/publish")
def publish_semester(data: schemas.SemesterPublish, db: Session = Depends(get_db)):
    db.query(models.Semester).update({models.Semester.is_current: False})

    semester = db.query(models.Semester).filter(models.Semester.name == data.name).first()
    if not semester:
        semester = models.Semester(
            name=data.name,
            start_date=data.start_date,
            end_date=data.end_date,
            is_enabled=data.is_enabled,
            is_current=True,
        )
        db.add(semester)
    else:
        semester.is_current = True
        semester.start_date = data.start_date
        semester.end_date = data.end_date
        semester.is_enabled = data.is_enabled

    db.commit()
    _sync_teacher_classes_from_students(db, semester.id)
    db.refresh(semester)
    return {

        "message": "当前学期已发布",
        "semester": {
            "id": semester.id,
            "name": semester.name,
            "start_date": semester.start_date,
            "end_date": semester.end_date,
            "is_enabled": semester.is_enabled,
            "is_current": semester.is_current,
        }
    }


@app.post("/admin/semesters/set-current")
def set_current_semester(data: schemas.SemesterSetCurrent, db: Session = Depends(get_db)):
    semester = db.query(models.Semester).filter(models.Semester.id == data.semester_id).first()
    if not semester:
        raise HTTPException(status_code=404, detail="学期不存在")

    db.query(models.Semester).update({models.Semester.is_current: False})
    semester.is_current = True
    db.commit()
    _sync_teacher_classes_from_students(db, semester.id)
    return {"message": "当前学期已切换"}



@app.post("/admin/classes")
def create_teaching_class(payload: schemas.TeachingClassCreate, db: Session = Depends(get_db)):
    canonical_teacher_id = _canonical_teacher_id(db, payload.teacher_id)
    if canonical_teacher_id is None:
        raise HTTPException(status_code=400, detail="teacher_id 无效")

    teacher = db.query(models.User).filter(models.User.id == canonical_teacher_id).first()
    if not teacher or teacher.role not in ["teacher", "admin"]:
        raise HTTPException(status_code=400, detail="teacher_id 无效")


    semester_id = payload.semester_id
    if semester_id is None:
        current = _get_current_semester(db)
        if not current:
            raise HTTPException(status_code=400, detail="请先发布当前学期")
        semester_id = current.id

    semester = db.query(models.Semester).filter(models.Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(status_code=404, detail="学期不存在")

    new_class = models.TeachingClass(name=payload.name, teacher_id=canonical_teacher_id, semester_id=semester_id)
    db.add(new_class)
    db.commit()
    db.refresh(new_class)
    return {"message": "班级创建成功", "class_id": new_class.id}


@app.post("/admin/classes/{class_id}/students")
def assign_students_to_class(class_id: int, data: schemas.ClassStudentAssign, db: Session = Depends(get_db)):
    teaching_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    canonical_teacher_id = _canonical_teacher_id(db, teaching_class.teacher_id)
    if canonical_teacher_id is None:
        raise HTTPException(status_code=400, detail="班级关联老师无效")
    teaching_class.teacher_id = canonical_teacher_id

    db.query(models.ClassStudent).filter(models.ClassStudent.class_id == class_id).delete()

    valid_students = db.query(models.User).filter(
        models.User.id.in_(data.student_ids),
        models.User.role == "student"
    ).all()

    for stu in valid_students:
        db.add(models.ClassStudent(class_id=class_id, student_id=stu.id))
        stu.class_name = teaching_class.name
        stu.teacher_id = canonical_teacher_id

    db.commit()
    return {"message": "班级学生名单已更新", "count": len(valid_students)}


@app.get("/admin/classes/manage")
def admin_manage_classes(
    semester_id: Optional[int] = Query(default=None),
    class_name: Optional[str] = Query(default=None),
    teacher_name: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    q = db.query(models.TeachingClass)
    if semester_id is not None:
        q = q.filter(models.TeachingClass.semester_id == semester_id)
    if class_name:
        q = q.filter(models.TeachingClass.name.ilike(f"%{class_name}%"))

    classes = q.order_by(models.TeachingClass.id.asc()).all()
    if not classes:
        return []

    semester_ids = {c.semester_id for c in classes}
    teacher_ids = {c.teacher_id for c in classes}

    semesters = db.query(models.Semester).filter(models.Semester.id.in_(list(semester_ids))).all() if semester_ids else []
    teachers = db.query(models.User).filter(models.User.id.in_(list(teacher_ids))).all() if teacher_ids else []

    semester_map = {s.id: s.name for s in semesters}
    teacher_map = {t.id: (t.real_name or t.username) for t in teachers}

    keyword_teacher = teacher_name.lower() if teacher_name else None
    result = []
    for item in classes:
        teacher_display = teacher_map.get(item.teacher_id)
        if keyword_teacher and (not teacher_display or keyword_teacher not in teacher_display.lower()):
            continue

        links = db.query(models.ClassStudent).filter(models.ClassStudent.class_id == item.id).all()
        student_ids = [x.student_id for x in links]
        students = db.query(models.User).filter(models.User.id.in_(student_ids)).all() if student_ids else []

        result.append({
            "id": item.id,
            "name": item.name,
            "semester_id": item.semester_id,
            "semester": semester_map.get(item.semester_id),
            "teacher_id": item.teacher_id,
            "teacher_name": teacher_display,
            "student_ids": student_ids,
            "students": [
                {
                    "id": s.id,
                    "username": s.username,
                    "real_name": s.real_name,
                    "student_no": s.student_no
                }
                for s in students
            ]
        })

    return result



@app.put("/admin/classes/{class_id}")
def update_teaching_class(class_id: int, payload: schemas.TeachingClassUpdate, db: Session = Depends(get_db)):
    item = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="班级不存在")

    if payload.teacher_id is not None:
        canonical_teacher_id = _canonical_teacher_id(db, payload.teacher_id)
        if canonical_teacher_id is None:
            raise HTTPException(status_code=400, detail="teacher_id 无效")
        item.teacher_id = canonical_teacher_id

    if payload.name is not None:
        name_text = payload.name.strip()
        if not name_text:
            raise HTTPException(status_code=400, detail="班级名称不能为空")
        item.name = name_text

    if payload.semester_id is not None:
        semester = db.query(models.Semester).filter(models.Semester.id == payload.semester_id).first()
        if not semester:
            raise HTTPException(status_code=404, detail="学期不存在")
        item.semester_id = payload.semester_id

    db.commit()
    return {"message": "班级信息更新成功"}



@app.delete("/admin/classes/{class_id}")
def delete_teaching_class(class_id: int, db: Session = Depends(get_db)):
    item = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="班级不存在")

    team_ids = [t.id for t in db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).all()]
    if team_ids:
        db.query(models.ClassTeamMember).filter(models.ClassTeamMember.team_id.in_(team_ids)).delete(synchronize_session=False)
    db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).delete(synchronize_session=False)
    db.query(models.ClassStudent).filter(models.ClassStudent.class_id == class_id).delete(synchronize_session=False)
    db.delete(item)
    db.commit()
    return {"message": "班级已删除"}


@app.get("/teachers/{teacher_id}/classes/current")
def get_teacher_current_classes(teacher_id: int, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id

    current = _get_current_semester(db)
    if not current:
        return []

    _sync_teacher_classes_from_students(db, current.id, teacher_id=teacher_id)

    classes = db.query(models.TeachingClass).filter(
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).order_by(models.TeachingClass.id.asc()).all()

    return [{"id": c.id, "name": c.name, "semester": current.name, "teacher_id": c.teacher_id} for c in classes]



@app.get("/admin/classes/current")
def get_all_current_classes(db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        return []

    _sync_teacher_classes_from_students(db, current.id)

    classes = db.query(models.TeachingClass).filter(models.TeachingClass.semester_id == current.id).order_by(models.TeachingClass.id.asc()).all()
    return [{"id": c.id, "name": c.name, "semester": current.name, "teacher_id": c.teacher_id} for c in classes]



@app.get("/students/{student_id}/classes/current")
def get_student_current_classes(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        return []

    _, changed = _ensure_student_class_relation(db, student, semester_id=current.id)
    if changed:
        db.commit()

    links = db.query(models.ClassStudent).filter(models.ClassStudent.student_id == student_id).all()
    class_ids = [x.class_id for x in links]
    if not class_ids:
        return []

    classes = db.query(models.TeachingClass).filter(
        models.TeachingClass.id.in_(class_ids),
        models.TeachingClass.semester_id == current.id
    ).order_by(models.TeachingClass.id.asc()).all()


    teacher_ids = {c.teacher_id for c in classes}
    teachers = db.query(models.User).filter(models.User.id.in_(list(teacher_ids))).all() if teacher_ids else []
    teacher_map = {t.id: (t.real_name or t.username) for t in teachers}

    return [
        {
            "id": c.id,
            "name": c.name,
            "semester": current.name,
            "semester_id": c.semester_id,
            "teacher_id": c.teacher_id,
            "teacher_name": teacher_map.get(c.teacher_id)
        }
        for c in classes
    ]


@app.get("/students/{student_id}/classes/{class_id}/assignments")
def get_student_class_assignments(student_id: int, class_id: int, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    relation = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == class_id,
        models.ClassStudent.student_id == student_id
    ).first()
    if not relation:
        raise HTTPException(status_code=403, detail="无权访问该班级作业")

    teacher = db.query(models.User).filter(models.User.id == target_class.teacher_id).first()
    assignments = db.query(models.Assignment).filter(
        (models.Assignment.course_id == class_id) | (models.Assignment.course_id.is_(None))
    ).order_by(models.Assignment.id.desc()).all()

    student_team_map, team_member_map = _get_class_team_member_map(db, class_id)
    my_team_id = student_team_map.get(student_id)
    my_team = db.query(models.ClassTeam).filter(models.ClassTeam.id == my_team_id).first() if my_team_id else None

    rows = []
    for item in assignments:
        target_type = _get_assignment_target_type(db, item.id, class_id=class_id)

        if target_type == "group":
            if not my_team_id:
                latest = None
            else:
                my_sub_ids = db.query(models.SubmissionTeamMeta.submission_id).filter(
                    models.SubmissionTeamMeta.team_id == my_team_id
                ).all()
                sub_ids = [x[0] for x in my_sub_ids]
                latest = db.query(models.Submission).filter(
                    models.Submission.assignment_id == item.id,
                    models.Submission.id.in_(sub_ids)
                ).order_by(models.Submission.id.desc()).first() if sub_ids else None
        else:
            latest = db.query(models.Submission).filter(
                models.Submission.assignment_id == item.id,
                models.Submission.student_id == student_id
            ).order_by(models.Submission.id.desc()).first()

        if not latest:
            status = "unsubmitted"
            status_label = "未提交"
        elif latest.status == "finished":
            status = "finished"
            status_label = "已批改"
        else:
            status = "submitted"
            status_label = "已提交"

        rows.append({
            "assignment_id": item.id,
            "title": item.title,
            "description": item.description,
            "deadline": item.deadline,
            "target_type": target_type,
            "team_id": my_team_id,
            "team_name": my_team.name if my_team else None,
            "status": status,
            "status_label": status_label,
            "score": latest.score if latest and latest.score is not None else None,
            "submission_id": latest.id if latest else None,
            "feedback": {
                "ai_criteria": (latest.ai_criteria if latest and latest.ai_criteria else item.ai_criteria),
                "ai_feedback": latest.ai_feedback if latest else None,
                "teacher_comment": latest.teacher_comment if latest else None,
                "score": latest.score if latest else None
            }
        })

    return {
        "class_id": target_class.id,
        "class_name": target_class.name,
        "semester": current.name,
        "teacher_id": target_class.teacher_id,
        "teacher_name": teacher.real_name if teacher and teacher.real_name else (teacher.username if teacher else None),
        "my_team": {
            "team_id": my_team_id,
            "team_name": my_team.name if my_team else None,
            "member_count": len(team_member_map.get(my_team_id, [])) if my_team_id else 0
        },
        "assignments": rows
    }



def _safe_file_part(text: Optional[str], fallback: str) -> str:
    raw = (text or "").strip()
    if not raw:
        raw = fallback
    return re.sub(r"[\\/:*?\"<>|\s]+", "_", raw)


@app.post("/students/{student_id}/classes/{class_id}/assignments/{assignment_id}/submit")
def submit_student_assignment_package(
    student_id: int,
    class_id: int,
    assignment_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    relation = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == class_id,
        models.ClassStudent.student_id == student_id
    ).first()
    if not relation:
        raise HTTPException(status_code=403, detail="无权提交该班级作业")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    assignment = db.query(models.Assignment).filter(models.Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="作业不存在")

    if assignment.course_id not in [None, class_id]:
        raise HTTPException(status_code=400, detail="该作业不属于当前班级")

    target_type = _get_assignment_target_type(db, assignment_id, class_id=class_id)
    student_team_map, _ = _get_class_team_member_map(db, class_id)
    team_id = student_team_map.get(student_id)

    if target_type == "group" and not team_id:
        raise HTTPException(status_code=400, detail="该作业为小组作业，你当前不在任何小组中")

    original_name = file.filename or "submission.zip"
    ext = Path(original_name).suffix.lower()
    allowed_ext = {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz"}
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail="仅支持上传 zip/rar/7z/tar/gz/bz2/xz 压缩文件")

    student_no = _safe_file_part(student.student_no, f"student_{student_id}")
    semester_text = _safe_file_part(current.name, "当前学期")
    class_text = _safe_file_part(target_class.name, f"class_{class_id}")
    assignment_text = _safe_file_part(assignment.title, f"assignment_{assignment_id}")

    if target_type == "group" and team_id:
        team = db.query(models.ClassTeam).filter(models.ClassTeam.id == team_id).first()
        team_text = _safe_file_part(team.name if team else f"team_{team_id}", f"team_{team_id}")
        base_dir = Path(__file__).resolve().parent / "codedoc" / f"team_{team_text}_{team_id}"
        final_name = f"{semester_text}_{class_text}_{assignment_text}_{team_text}{ext}"
    else:
        base_dir = Path(__file__).resolve().parent / "codedoc" / student_no
        final_name = f"{semester_text}_{class_text}_{assignment_text}{ext}"

    base_dir.mkdir(parents=True, exist_ok=True)
    final_path = base_dir / final_name

    with final_path.open("wb") as f:
        f.write(file.file.read())

    submission = models.Submission(
        assignment_id=assignment_id,
        student_id=student_id,
        code_content=f"[压缩包提交]{final_path.as_posix()}",
        status="submitted"
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)

    db.add(models.SubmissionTeamMeta(
        submission_id=submission.id,
        team_id=team_id if target_type == "group" else None,
        submitter_id=student_id
    ))
    db.commit()

    return {
        "message": "作业压缩包上传成功",
        "submission_id": submission.id,
        "saved_path": final_path.as_posix(),
        "file_name": final_name,
        "target_type": target_type,
        "team_id": team_id if target_type == "group" else None
    }



@app.post("/students/{student_id}/classes/{class_id}/report-tasks/{report_task_id}/submit")
def submit_student_report_document(
    student_id: int,
    class_id: int,
    report_task_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    relation = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == class_id,
        models.ClassStudent.student_id == student_id
    ).first()
    if not relation:
        raise HTTPException(status_code=403, detail="无权提交该班级报告")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    report_task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_task_id,
        models.ReportTask.class_id == class_id
    ).first()
    if not report_task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    target_type = _get_report_task_target_type(db, report_task_id, class_id=class_id)
    student_team_map, _ = _get_class_team_member_map(db, class_id)
    team_id = student_team_map.get(student_id)

    if target_type == "group" and not team_id:
        raise HTTPException(status_code=400, detail="该报告为小组任务，你当前不在任何小组中")

    original_name = file.filename or "report.docx"
    ext = Path(original_name).suffix.lower()
    allowed_ext = {".doc", ".docx", ".pdf"}
    if ext not in allowed_ext:
        raise HTTPException(status_code=400, detail="仅支持上传 doc/docx/pdf 文件")

    latest_submission = None
    if target_type == "group" and team_id:
        team_sub_ids = db.query(models.ReportSubmissionTeamMeta.report_submission_id).filter(
            models.ReportSubmissionTeamMeta.team_id == team_id
        ).all()
        sub_ids = [x[0] for x in team_sub_ids]
        latest_submission = db.query(models.ReportSubmission).filter(
            models.ReportSubmission.report_task_id == report_task_id,
            models.ReportSubmission.id.in_(sub_ids)
        ).order_by(models.ReportSubmission.id.desc()).first() if sub_ids else None
    else:
        latest_submission = db.query(models.ReportSubmission).filter(
            models.ReportSubmission.report_task_id == report_task_id,
            models.ReportSubmission.student_id == student_id
        ).order_by(models.ReportSubmission.id.desc()).first()

    if latest_submission and latest_submission.status == "finished":
        raise HTTPException(status_code=400, detail="该报告已被老师批改，若需修改请联系老师先退回")

    student_no = _safe_file_part(student.student_no, f"student_{student_id}")
    class_text = _safe_file_part(target_class.name, f"class_{class_id}")
    student_name_text = _safe_file_part(student.real_name or student.username, f"student_{student_id}")
    task_text = _safe_file_part(report_task.title, f"report_task_{report_task_id}")

    if target_type == "group" and team_id:
        team = db.query(models.ClassTeam).filter(models.ClassTeam.id == team_id).first()
        team_text = _safe_file_part(team.name if team else f"team_{team_id}", f"team_{team_id}")
        base_dir = Path(__file__).resolve().parent / "document" / f"team_{team_text}_{team_id}"
        final_name = f"{class_text}_{team_text}_{task_text}{ext}"
    else:
        base_dir = Path(__file__).resolve().parent / "document" / student_no
        final_name = f"{class_text}_{student_name_text}_{task_text}{ext}"

    base_dir.mkdir(parents=True, exist_ok=True)
    final_path = base_dir / final_name

    with final_path.open("wb") as f:
        f.write(file.file.read())

    submission = models.ReportSubmission(
        report_task_id=report_task_id,
        student_id=student_id,
        file_path=final_path.as_posix(),
        original_filename=original_name,
        status="submitted"
    )

    db.add(submission)
    db.commit()
    db.refresh(submission)

    db.add(models.ReportSubmissionTeamMeta(
        report_submission_id=submission.id,
        team_id=team_id if target_type == "group" else None,
        submitter_id=student_id
    ))
    db.commit()

    return {
        "message": "报告上传成功",
        "submission_id": submission.id,
        "saved_path": final_path.as_posix(),
        "file_name": final_name,
        "target_type": target_type,
        "team_id": team_id if target_type == "group" else None
    }



def _get_students_for_class(db: Session, teaching_class: models.TeachingClass):
    canonical_teacher_id = _canonical_teacher_id(db, teaching_class.teacher_id) or teaching_class.teacher_id

    class_student_rows = db.query(models.ClassStudent).filter(models.ClassStudent.class_id == teaching_class.id).all()
    student_ids = [item.student_id for item in class_student_rows]

    if student_ids:
        return db.query(models.User).filter(
            models.User.id.in_(student_ids),
            models.User.role == "student"
        ).all()

    return db.query(models.User).filter(
        models.User.role == "student",
        models.User.teacher_id == canonical_teacher_id,
        models.User.class_name == teaching_class.name
    ).all()


def _get_teams_for_class(db: Session, class_id: int):
    teams = db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).all()
    result = []
    for team in teams:
        members = db.query(models.ClassTeamMember).filter(models.ClassTeamMember.team_id == team.id).all()
        result.append({"id": team.id, "name": team.name, "members": [m.student_id for m in members]})
    return result



def _normalize_target_type(raw: Optional[str]) -> str:
    text = (raw or "individual").strip().lower()
    return "group" if text == "group" else "individual"



def _get_assignment_target_type(db: Session, assignment_id: int, class_id: Optional[int] = None) -> str:
    query = db.query(models.AssignmentPublishConfig).filter(models.AssignmentPublishConfig.assignment_id == assignment_id)
    if class_id is not None:
        query = query.filter(models.AssignmentPublishConfig.class_id == class_id)
    config = query.first()
    return _normalize_target_type(config.target_type if config else "individual")



def _get_report_task_target_type(db: Session, report_task_id: int, class_id: Optional[int] = None) -> str:
    query = db.query(models.ReportTaskPublishConfig).filter(models.ReportTaskPublishConfig.report_task_id == report_task_id)
    if class_id is not None:
        query = query.filter(models.ReportTaskPublishConfig.class_id == class_id)
    config = query.first()
    return _normalize_target_type(config.target_type if config else "individual")



def _get_class_team_member_map(db: Session, class_id: int):
    teams = db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).all()
    if not teams:
        return {}, {}

    team_ids = [x.id for x in teams]
    members = db.query(models.ClassTeamMember).filter(models.ClassTeamMember.team_id.in_(team_ids)).all()

    student_team_map = {}
    team_member_map = {team.id: [] for team in teams}
    for item in members:
        student_team_map[item.student_id] = item.team_id
        if item.team_id in team_member_map:
            team_member_map[item.team_id].append(item.student_id)

    return student_team_map, team_member_map



@app.get("/teachers/{teacher_id}/classes/{class_id}/dashboard")
def get_teacher_class_dashboard(teacher_id: int, class_id: int, db: Session = Depends(get_db)):

    teacher = _require_teacher_user(db, teacher_id)
    teacher_id = teacher.id

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    teaching_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="未找到当前学期下该老师的班级")

    students = _get_students_for_class(db, teaching_class)
    assignments = db.query(models.Assignment).filter(
        models.Assignment.course_id == class_id
    ).order_by(models.Assignment.id.asc()).all()
    teams = _get_teams_for_class(db, class_id)


    student_list = []
    for stu in students:
        scores = {}
        for assign in assignments:
            sub = db.query(models.Submission).filter(
                models.Submission.student_id == stu.id,
                models.Submission.assignment_id == assign.id,
                models.Submission.status == "finished"
            ).order_by(models.Submission.id.desc()).first()
            scores[str(assign.id)] = sub.score if sub and sub.score is not None else None

        student_list.append({
            "id": stu.id,
            "student_no": stu.student_no,
            "username": stu.username,
            "name": stu.real_name or stu.username,
            "class_name": stu.class_name,
            "scores": scores
        })

    return {
        "class_id": teaching_class.id,
        "class_name": teaching_class.name,
        "semester": current.name,
        "assignments": [{"id": a.id, "title": a.title} for a in assignments],
        "students": student_list,
        "teams": teams,
    }


def _build_assignment_stats_for_class(db: Session, class_id: int, class_students: list[models.User]):
    student_ids = [s.id for s in class_students]
    assignments = db.query(models.Assignment).filter(
        models.Assignment.course_id == class_id
    ).order_by(models.Assignment.id.desc()).all()
    result = []


    student_team_map, team_member_map = _get_class_team_member_map(db, class_id)
    group_team_ids = [team_id for team_id, member_ids in team_member_map.items() if member_ids]

    for item in assignments:
        target_type = _get_assignment_target_type(db, item.id, class_id=class_id)
        subs = db.query(models.Submission).filter(models.Submission.assignment_id == item.id).all()
        if student_ids:
            subs = [s for s in subs if s.student_id in student_ids]

        if target_type == "group":
            meta_rows = db.query(models.SubmissionTeamMeta).filter(
                models.SubmissionTeamMeta.submission_id.in_([s.id for s in subs])
            ).all() if subs else []
            submission_team_map = {m.submission_id: m.team_id for m in meta_rows if m.team_id}
            submitted_team_ids = {submission_team_map.get(s.id) for s in subs if submission_team_map.get(s.id) in group_team_ids}
            pending_review = sum(1 for s in subs if submission_team_map.get(s.id) in group_team_ids and s.status != "finished")
            unit_total = len(group_team_ids)
            unsubmitted = max(unit_total - len(submitted_team_ids), 0)
        else:
            submitted_student_ids = {s.student_id for s in subs}
            pending_review = sum(1 for s in subs if s.status != "finished")
            unit_total = len(student_ids)
            unsubmitted = max(unit_total - len(submitted_student_ids), 0)

        result.append({
            "id": item.id,
            "title": item.title,
            "description": item.description,
            "deadline": item.deadline,
            "ai_criteria": item.ai_criteria,
            "target_type": target_type,
            "pending_review_count": pending_review,
            "unsubmitted_count": unsubmitted,
            "student_total": len(student_ids),
            "unit_total": unit_total,
        })

    return result



@app.get("/teachers/{teacher_id}/classes/{class_id}/assignments")
def get_teacher_class_assignments(teacher_id: int, class_id: int, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在或不属于当前老师")

    students = _get_students_for_class(db, target_class)
    assignments = _build_assignment_stats_for_class(db, class_id, students)

    return {
        "class_id": target_class.id,
        "class_name": target_class.name,
        "semester": current.name,
        "students": [{"id": s.id, "name": s.real_name or s.username, "username": s.username} for s in students],
        "assignments": assignments
    }


def _build_report_task_stats_for_class(db: Session, class_id: int, class_students: list[models.User]):
    student_ids = [s.id for s in class_students]
    tasks = db.query(models.ReportTask).filter(
        models.ReportTask.class_id == class_id
    ).order_by(models.ReportTask.id.desc()).all()

    student_team_map, team_member_map = _get_class_team_member_map(db, class_id)
    group_team_ids = [team_id for team_id, member_ids in team_member_map.items() if member_ids]

    result = []
    for item in tasks:
        target_type = _get_report_task_target_type(db, item.id, class_id=class_id)
        subs = db.query(models.ReportSubmission).filter(
            models.ReportSubmission.report_task_id == item.id
        ).all()
        if student_ids:
            subs = [s for s in subs if s.student_id in student_ids]

        if target_type == "group":
            meta_rows = db.query(models.ReportSubmissionTeamMeta).filter(
                models.ReportSubmissionTeamMeta.report_submission_id.in_([s.id for s in subs])
            ).all() if subs else []
            submission_team_map = {m.report_submission_id: m.team_id for m in meta_rows if m.team_id}

            latest_status_by_team: dict[int, str] = {}
            for sub in sorted(subs, key=lambda x: x.id, reverse=True):
                sub_team_id = submission_team_map.get(sub.id)
                if not sub_team_id or sub_team_id not in group_team_ids:
                    continue
                if sub_team_id not in latest_status_by_team:
                    latest_status_by_team[sub_team_id] = sub.status

            pending_review = sum(1 for status in latest_status_by_team.values() if status == "submitted")
            submitted_team_ids = {
                team_id for team_id, status in latest_status_by_team.items()
                if status in ["submitted", "finished"]
            }
            unit_total = len(group_team_ids)
            unsubmitted = max(unit_total - len(submitted_team_ids), 0)
        else:
            latest_status_by_student: dict[int, str] = {}
            for sub in sorted(subs, key=lambda x: x.id, reverse=True):
                if sub.student_id not in latest_status_by_student:
                    latest_status_by_student[sub.student_id] = sub.status

            pending_review = sum(1 for status in latest_status_by_student.values() if status == "submitted")
            submitted_student_ids = {
                sid for sid, status in latest_status_by_student.items()
                if status in ["submitted", "finished"]
            }
            unit_total = len(student_ids)
            unsubmitted = max(unit_total - len(submitted_student_ids), 0)


        result.append({
            "id": item.id,
            "title": item.title,
            "description": item.description,
            "deadline": item.deadline,
            "target_type": target_type,
            "pending_review_count": pending_review,
            "unsubmitted_count": unsubmitted,
            "student_total": len(student_ids),
            "unit_total": unit_total,
            "created_at": item.created_at.isoformat() if item.created_at else None
        })

    return result



@app.get("/teachers/{teacher_id}/classes/{class_id}/report-tasks")
def get_teacher_class_report_tasks(teacher_id: int, class_id: int, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在或不属于当前老师")

    students = _get_students_for_class(db, target_class)
    tasks = _build_report_task_stats_for_class(db, class_id, students)

    return {
        "class_id": target_class.id,
        "class_name": target_class.name,
        "semester": current.name,
        "students": [{"id": s.id, "name": s.real_name or s.username, "username": s.username} for s in students],
        "report_tasks": tasks
    }


@app.get("/teachers/{teacher_id}/classes/{class_id}/report-tasks/{report_task_id}/submissions")
def get_teacher_report_task_submissions(
    teacher_id: int,
    class_id: int,
    report_task_id: int,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在或不属于当前老师")

    task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_task_id,
        models.ReportTask.class_id == class_id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    target_type = _get_report_task_target_type(db, report_task_id, class_id=class_id)
    students = _get_students_for_class(db, target_class)
    student_ids = [s.id for s in students]

    if target_type == "group":
        teams = db.query(models.ClassTeam).filter(models.ClassTeam.class_id == class_id).all()
        team_map = {t.id: t for t in teams}

        member_rows = db.query(models.ClassTeamMember).filter(
            models.ClassTeamMember.team_id.in_(list(team_map.keys()))
        ).all() if team_map else []
        member_student_ids = [x.student_id for x in member_rows]
        member_users = db.query(models.User).filter(models.User.id.in_(member_student_ids)).all() if member_student_ids else []
        member_user_map = {u.id: u for u in member_users}

        team_member_names: dict[int, list[str]] = {tid: [] for tid in team_map.keys()}
        for m in member_rows:
            user = member_user_map.get(m.student_id)
            if not user:
                continue
            team_member_names[m.team_id].append(user.real_name or user.username)

        all_subs = db.query(models.ReportSubmission).filter(
            models.ReportSubmission.report_task_id == report_task_id,
            models.ReportSubmission.student_id.in_(student_ids)
        ).order_by(models.ReportSubmission.id.desc()).all() if student_ids else []

        sub_ids = [s.id for s in all_subs]
        meta_rows = db.query(models.ReportSubmissionTeamMeta).filter(
            models.ReportSubmissionTeamMeta.report_submission_id.in_(sub_ids)
        ).all() if sub_ids else []
        submission_team_map = {m.report_submission_id: m.team_id for m in meta_rows if m.team_id}

        latest_team_sub: dict[int, models.ReportSubmission] = {}
        for sub in all_subs:
            team_id = submission_team_map.get(sub.id)
            if team_id and (team_id not in latest_team_sub):
                latest_team_sub[team_id] = sub

        rows = []
        for team_id, team in team_map.items():
            if not team_member_names.get(team_id):
                continue

            latest = latest_team_sub.get(team_id)
            if not latest:
                status = "unsubmitted"
                status_label = "未提交"
            elif latest.status == "finished":
                status = "finished"
                status_label = "已批改"
            elif latest.status == "returned":
                status = "returned"
                status_label = "已退回"
            else:
                status = "submitted"
                status_label = "已提交"


            rows.append({
                "student_id": -team_id,
                "student_name": f"{team.name}（小组）",
                "student_no": f"{len(team_member_names.get(team_id, []))}人",
                "team_id": team_id,
                "team_name": team.name,
                "team_members": team_member_names.get(team_id, []),
                "status": status,
                "status_label": status_label,
                "submission_id": latest.id if latest else None,
                "score": latest.score if latest and latest.score is not None else None,
                "saved_path": latest.file_path if latest else None,
                "feedback": {
                    "ai_feedback": latest.ai_feedback if latest else None,
                    "teacher_comment": latest.teacher_comment if latest else None,
                    "score": latest.score if latest else None,
                    "ai_criteria": latest.ai_criteria if latest else None
                }
            })

        return {
            "class_id": class_id,
            "class_name": target_class.name,
            "report_task_id": task.id,
            "report_task_title": task.title,
            "target_type": "group",
            "students": rows
        }

    latest_sub_map: dict[int, models.ReportSubmission] = {}
    if student_ids:
        all_subs = db.query(models.ReportSubmission).filter(
            models.ReportSubmission.report_task_id == report_task_id,
            models.ReportSubmission.student_id.in_(student_ids)
        ).order_by(models.ReportSubmission.id.desc()).all()

        for sub in all_subs:
            if sub.student_id not in latest_sub_map:
                latest_sub_map[sub.student_id] = sub

    rows = []
    for stu in students:
        latest = latest_sub_map.get(stu.id)
        if not latest:
            status = "unsubmitted"
            status_label = "未提交"
        elif latest.status == "finished":
            status = "finished"
            status_label = "已批改"
        elif latest.status == "returned":
            status = "returned"
            status_label = "已退回"
        else:
            status = "submitted"
            status_label = "已提交"

        rows.append({

            "student_id": stu.id,
            "student_name": stu.real_name or stu.username,
            "student_no": stu.student_no,
            "status": status,
            "status_label": status_label,
            "submission_id": latest.id if latest else None,
            "score": latest.score if latest and latest.score is not None else None,
            "saved_path": latest.file_path if latest else None,
            "feedback": {
                "ai_feedback": latest.ai_feedback if latest else None,
                "teacher_comment": latest.teacher_comment if latest else None,
                "score": latest.score if latest else None,
                "ai_criteria": latest.ai_criteria if latest else None
            }
        })

    return {
        "class_id": class_id,
        "class_name": target_class.name,
        "report_task_id": task.id,
        "report_task_title": task.title,
        "target_type": "individual",
        "students": rows
    }



@app.post("/teachers/{teacher_id}/report-submissions/{report_submission_id}/publish-report")
def publish_report_submission_review(
    teacher_id: int,
    report_submission_id: int,
    payload: schemas.ReportSubmissionReviewCreate,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    report_submission = db.query(models.ReportSubmission).filter(
        models.ReportSubmission.id == report_submission_id
    ).first()
    if not report_submission:
        raise HTTPException(status_code=404, detail="报告提交记录不存在")

    report_task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_submission.report_task_id
    ).first()
    if not report_task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == report_task.class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=403, detail="无权发布该报告批改结果")

    if report_submission.status == "returned":
        raise HTTPException(status_code=400, detail="该报告已退回，需学生重新提交后才能发布批改结果")

    report_submission.ai_feedback = payload.ai_feedback

    report_submission.teacher_comment = payload.teacher_comment
    report_submission.score = payload.score
    report_submission.ai_criteria = payload.ai_criteria
    report_submission.status = "finished"

    db.commit()

    return {
        "message": "报告批改结果已发布",
        "report_submission_id": report_submission.id,
        "status": report_submission.status
    }


@app.post("/teachers/{teacher_id}/report-submissions/{report_submission_id}/return")
def return_report_submission(
    teacher_id: int,
    report_submission_id: int,
    payload: schemas.ReportSubmissionReturnCreate,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    report_submission = db.query(models.ReportSubmission).filter(
        models.ReportSubmission.id == report_submission_id
    ).first()
    if not report_submission:
        raise HTTPException(status_code=404, detail="报告提交记录不存在")

    report_task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_submission.report_task_id
    ).first()
    if not report_task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == report_task.class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=403, detail="无权退回该报告")

    removed_file = False
    saved_path = (report_submission.file_path or "").strip()
    if saved_path:
        report_file = Path(saved_path)
        try:
            if report_file.exists() and report_file.is_file():
                report_file.unlink()
                removed_file = True
        except Exception:
            removed_file = False

    reason = (payload.reason or "").strip()
    if reason:
        base_comment = (report_submission.teacher_comment or "").strip()
        return_note = f"【退回原因】{reason}"
        report_submission.teacher_comment = f"{base_comment}\n{return_note}".strip() if base_comment else return_note

    report_submission.status = "returned"
    report_submission.file_path = ""
    db.commit()

    return {
        "message": "报告已退回，学生可重新提交",
        "report_submission_id": report_submission.id,
        "status": report_submission.status,
        "file_deleted": removed_file
    }


@app.get("/teachers/{teacher_id}/report-submissions/{report_submission_id}/ai-history")

def get_report_ai_history(
    teacher_id: int,
    report_submission_id: int,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    report_submission = db.query(models.ReportSubmission).filter(
        models.ReportSubmission.id == report_submission_id
    ).first()
    if not report_submission:
        raise HTTPException(status_code=404, detail="报告提交记录不存在")

    report_task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_submission.report_task_id
    ).first()
    if not report_task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == report_task.class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=403, detail="无权查看该报告AI历史")

    md_path = _build_report_student_history_path(db, report_submission)
    records = _parse_report_ai_history_records(md_path)

    return {
        "report_submission_id": report_submission.id,
        "student_id": report_submission.student_id,
        "history_file": md_path.as_posix(),
        "records": records,
    }


@app.delete("/teachers/{teacher_id}/report-submissions/{report_submission_id}/ai-history/{record_id}")
def delete_report_ai_history_record(
    teacher_id: int,
    report_submission_id: int,
    record_id: int,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    report_submission = db.query(models.ReportSubmission).filter(
        models.ReportSubmission.id == report_submission_id
    ).first()
    if not report_submission:
        raise HTTPException(status_code=404, detail="报告提交记录不存在")

    report_task = db.query(models.ReportTask).filter(
        models.ReportTask.id == report_submission.report_task_id
    ).first()
    if not report_task:
        raise HTTPException(status_code=404, detail="报告任务不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == report_task.class_id,
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=403, detail="无权删除该报告AI历史")

    md_path = _build_report_student_history_path(db, report_submission)
    deleted = _delete_report_ai_history_record(md_path, record_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="指定历史记录不存在")

    return {
        "message": "AI评测历史记录已删除",
        "record_id": record_id,
        "history_file": md_path.as_posix(),
    }


@app.get("/students/{student_id}/classes/{class_id}/report-tasks")
def get_student_class_report_tasks(student_id: int, class_id: int, db: Session = Depends(get_db)):


    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    target_class = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not target_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    relation = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == class_id,
        models.ClassStudent.student_id == student_id
    ).first()
    if not relation:
        raise HTTPException(status_code=403, detail="无权访问该班级报告任务")

    teacher = db.query(models.User).filter(models.User.id == target_class.teacher_id).first()
    tasks = db.query(models.ReportTask).filter(
        models.ReportTask.class_id == class_id
    ).order_by(models.ReportTask.id.desc()).all()

    student_team_map, team_member_map = _get_class_team_member_map(db, class_id)
    my_team_id = student_team_map.get(student_id)
    my_team = db.query(models.ClassTeam).filter(models.ClassTeam.id == my_team_id).first() if my_team_id else None

    rows = []
    for item in tasks:
        target_type = _get_report_task_target_type(db, item.id, class_id=class_id)

        if target_type == "group":
            if not my_team_id:
                latest = None
            else:
                my_sub_ids = db.query(models.ReportSubmissionTeamMeta.report_submission_id).filter(
                    models.ReportSubmissionTeamMeta.team_id == my_team_id
                ).all()
                sub_ids = [x[0] for x in my_sub_ids]
                latest = db.query(models.ReportSubmission).filter(
                    models.ReportSubmission.report_task_id == item.id,
                    models.ReportSubmission.id.in_(sub_ids)
                ).order_by(models.ReportSubmission.id.desc()).first() if sub_ids else None
        else:
            latest = db.query(models.ReportSubmission).filter(
                models.ReportSubmission.report_task_id == item.id,
                models.ReportSubmission.student_id == student_id
            ).order_by(models.ReportSubmission.id.desc()).first()

        if not latest:
            status = "unsubmitted"
            status_label = "未提交"
        elif latest.status == "finished":
            status = "finished"
            status_label = "已批改"
        elif latest.status == "returned":
            status = "returned"
            status_label = "已退回"
        else:
            status = "submitted"
            status_label = "已提交"

        rows.append({

            "report_task_id": item.id,
            "title": item.title,
            "description": item.description,
            "deadline": item.deadline,
            "target_type": target_type,
            "team_id": my_team_id,
            "team_name": my_team.name if my_team else None,
            "status": status,
            "status_label": status_label,
            "score": latest.score if latest and latest.score is not None else None,
            "submission_id": latest.id if latest else None,
            "saved_path": latest.file_path if latest else None,
            "feedback": {
                "ai_feedback": latest.ai_feedback if latest else None,
                "teacher_comment": latest.teacher_comment if latest else None,
                "score": latest.score if latest else None,
                "ai_criteria": latest.ai_criteria if latest else None
            }
        })

    return {
        "class_id": target_class.id,
        "class_name": target_class.name,
        "semester": current.name,
        "teacher_id": target_class.teacher_id,
        "teacher_name": teacher.real_name if teacher and teacher.real_name else (teacher.username if teacher else None),
        "my_team": {
            "team_id": my_team_id,
            "team_name": my_team.name if my_team else None,
            "member_count": len(team_member_map.get(my_team_id, [])) if my_team_id else 0
        },
        "report_tasks": rows
    }



@app.get("/teachers/{teacher_id}/dashboard/summary")
def get_teacher_dashboard_summary(teacher_id: int, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id

    current = _get_current_semester(db)
    if not current:
        return {"semester": None, "class_count": 0, "student_count": 0, "assignment_count": 0, "pending_review_count": 0}

    _sync_teacher_classes_from_students(db, current.id, teacher_id=teacher_id)

    classes = db.query(models.TeachingClass).filter(
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).all()


    class_ids = [c.id for c in classes]
    class_students = []
    for c in classes:
        class_students.extend(_get_students_for_class(db, c))

    unique_students = {s.id: s for s in class_students}
    assignments = db.query(models.Assignment).filter(models.Assignment.course_id.in_(class_ids)).all() if class_ids else []

    pending_review_count = 0
    student_id_set = set(unique_students.keys())
    for a in assignments:
        subs = db.query(models.Submission).filter(models.Submission.assignment_id == a.id).all()
        pending_review_count += sum(1 for s in subs if s.student_id in student_id_set and s.status != "finished")

    return {
        "semester": current.name,
        "class_count": len(classes),
        "student_count": len(unique_students),
        "assignment_count": len(assignments),
        "pending_review_count": pending_review_count
    }


def _parse_deadline_date(deadline: Optional[str]):
    if not deadline:
        return None

    text = str(deadline).strip()
    if not text:
        return None

    fmts = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"]
    for fmt in fmts:
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


@app.get("/teachers/{teacher_id}/dashboard/trend")
def get_teacher_dashboard_trend(
    teacher_id: int,
    mode: str = Query(default="day"),
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id

    mode = (mode or "day").lower()
    if mode not in ["day", "week"]:
        raise HTTPException(status_code=400, detail="mode 仅支持 day/week")

    current = _get_current_semester(db)
    if not current:
        labels = [f"D{i}" for i in range(1, 8)] if mode == "day" else [f"W{i}" for i in range(1, 9)]
        return {
            "mode": mode,
            "labels": labels,
            "series": [
                {"name": "发布作业", "data": [0 for _ in labels]},
                {"name": "提交人数", "data": [0 for _ in labels]},
                {"name": "已批改人数", "data": [0 for _ in labels]},
            ]
        }

    _sync_teacher_classes_from_students(db, current.id, teacher_id=teacher_id)

    classes = db.query(models.TeachingClass).filter(
        models.TeachingClass.teacher_id == teacher_id,
        models.TeachingClass.semester_id == current.id
    ).all()

    class_ids = [c.id for c in classes]

    students = []
    for c in classes:
        students.extend(_get_students_for_class(db, c))
    student_id_set = {s.id for s in students}

    assignments = db.query(models.Assignment).filter(
        (models.Assignment.course_id.in_(class_ids)) | (models.Assignment.course_id.is_(None))
    ).all() if class_ids else db.query(models.Assignment).filter(models.Assignment.course_id.is_(None)).all()
    assignment_ids = [a.id for a in assignments]
    submissions = db.query(models.Submission).filter(models.Submission.assignment_id.in_(assignment_ids)).all() if assignment_ids else []
    if student_id_set:
        submissions = [s for s in submissions if s.student_id in student_id_set]

    subs_by_assignment: dict[int, list] = {}
    for s in submissions:
        subs_by_assignment.setdefault(s.assignment_id, []).append(s)

    today = datetime.now().date()

    if mode == "day":
        day_points = [today + timedelta(days=i) for i in range(-7, 15)]
        labels = [d.strftime("%m-%d") for d in day_points]
        idx = {d: i for i, d in enumerate(day_points)}
    else:
        week_start = today - timedelta(days=today.weekday())
        week_points = [week_start + timedelta(weeks=i) for i in range(-4, 5)]
        labels = [f"W{d.isocalendar().week:02d}" for d in week_points]
        idx = {(d.isocalendar().year, d.isocalendar().week): i for i, d in enumerate(week_points)}

    published = [0 for _ in labels]
    submitted = [0 for _ in labels]
    finished = [0 for _ in labels]

    for a in assignments:
        deadline_date = _parse_deadline_date(a.deadline)
        if not deadline_date:
            continue

        if mode == "day":
            i = idx.get(deadline_date)
        else:
            i = idx.get((deadline_date.isocalendar().year, deadline_date.isocalendar().week))

        if i is None:
            continue

        published[i] += 1
        class_subs = subs_by_assignment.get(a.id, [])
        submitted[i] += len({x.student_id for x in class_subs})
        finished[i] += len({x.student_id for x in class_subs if x.status == "finished"})

    return {
        "mode": mode,
        "labels": labels,
        "series": [
            {"name": "发布作业", "data": published},
            {"name": "提交人数", "data": submitted},
            {"name": "已批改人数", "data": finished},
        ]
    }


@app.get("/classes/{class_id}/analytics/student_ranking")
def get_class_student_ranking(class_id: int, db: Session = Depends(get_db)):
    teaching_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    students = _get_students_for_class(db, teaching_class)
    if not students:
        return []

    student_ids = [s.id for s in students]
    name_map = {s.id: (s.real_name or s.username) for s in students}

    assignments = db.query(models.Assignment).filter(
        (models.Assignment.course_id == class_id) | (models.Assignment.course_id.is_(None))
    ).all()
    assignment_ids = [a.id for a in assignments]
    if not assignment_ids:
        return []

    subs = db.query(models.Submission).filter(
        models.Submission.assignment_id.in_(assignment_ids),
        models.Submission.student_id.in_(student_ids),
        models.Submission.status == "finished"
    ).all()

    stat: dict[int, list] = {}
    for s in subs:
        if s.score is None:
            continue
        stat.setdefault(s.student_id, []).append(float(s.score))

    rows = []
    for sid, scores in stat.items():
        if not scores:
            continue
        avg = round(sum(scores) / len(scores), 2)
        rows.append({"name": name_map.get(sid, f"学生{sid}"), "score": avg, "count": len(scores)})

    rows.sort(key=lambda x: x["score"], reverse=True)
    return rows


@app.get("/classes/{class_id}/overview")
def get_class_overview(class_id: int, db: Session = Depends(get_db)):
    teaching_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    students = _get_students_for_class(db, teaching_class)
    assignments = db.query(models.Assignment).filter(
        models.Assignment.course_id == class_id
    ).order_by(models.Assignment.id.asc()).all()

    student_list = []

    for stu in students:
        scores = {}
        for assign in assignments:
            sub = db.query(models.Submission).filter(
                models.Submission.student_id == stu.id,
                models.Submission.assignment_id == assign.id,
                models.Submission.status == "finished"
            ).order_by(models.Submission.id.desc()).first()
            scores[str(assign.id)] = sub.score if sub and sub.score is not None else None

        student_list.append({
            "id": stu.id,
            "student_no": stu.student_no,
            "username": stu.username,
            "name": stu.real_name or stu.username,
            "class_name": stu.class_name,
            "scores": scores
        })

    return {
        "class_id": teaching_class.id,
        "class_name": teaching_class.name,
        "assignments": [{"id": a.id, "title": a.title} for a in assignments],
        "students": student_list
    }


@app.get("/classes/{class_id}/teams")
def get_class_teams(class_id: int, db: Session = Depends(get_db)):
    teaching_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    return _get_teams_for_class(db, class_id)


@app.get("/students/{student_id}/teams")
def get_student_team_overview(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    links = db.query(models.ClassStudent).filter(models.ClassStudent.student_id == student_id).all()
    class_ids = [x.class_id for x in links]
    if not class_ids:
        return []

    classes = db.query(models.TeachingClass).filter(models.TeachingClass.id.in_(class_ids)).all()
    result = []
    for item in classes:
        teams = _get_teams_for_class(db, item.id)

        my_team = None
        for t in teams:
            if student_id in (t.get("members") or []):
                my_team = t
                break

        member_ids = my_team.get("members", []) if my_team else []
        users = db.query(models.User).filter(models.User.id.in_(member_ids)).all() if member_ids else []
        members = [
            {
                "id": u.id,
                "name": u.real_name or u.username,
                "student_no": u.student_no,
                "username": u.username,
            }
            for u in users
        ]

        result.append({
            "class_id": item.id,
            "class_name": item.name,
            "team": {
                "id": my_team.get("id") if my_team else None,
                "name": my_team.get("name") if my_team else None,
                "members": members,
            },
            "team_total": len(teams),
        })

    return result


@app.post("/classes/{class_id}/teams")
def create_class_team(class_id: int, data: schemas.ClassTeamCreate, db: Session = Depends(get_db)):

    teaching_class = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not teaching_class:
        raise HTTPException(status_code=404, detail="班级不存在")

    team = models.ClassTeam(class_id=class_id, name=data.name)
    db.add(team)
    db.commit()
    db.refresh(team)
    return {"message": "队伍创建成功", "team_id": team.id}


@app.put("/class-teams/{team_id}/members")
def update_team_members(team_id: int, data: schemas.TeamMembersUpdate, db: Session = Depends(get_db)):
    team = db.query(models.ClassTeam).filter(models.ClassTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")

    class_students = db.query(models.ClassStudent).filter(models.ClassStudent.class_id == team.class_id).all()
    valid_student_ids = {item.student_id for item in class_students}
    selected_ids = [sid for sid in data.student_ids if sid in valid_student_ids]

    db.query(models.ClassTeamMember).filter(models.ClassTeamMember.team_id == team_id).delete()

    for sid in selected_ids:
        db.query(models.ClassTeamMember).filter(models.ClassTeamMember.student_id == sid).delete()
        db.add(models.ClassTeamMember(team_id=team_id, student_id=sid))

    db.commit()
    return {"message": "队伍成员已更新", "count": len(selected_ids)}


@app.post("/classes/{class_id}/transfer-member")
def transfer_team_member(class_id: int, data: schemas.TeamTransfer, db: Session = Depends(get_db)):
    target_team = db.query(models.ClassTeam).filter(
        models.ClassTeam.id == data.target_team_id,
        models.ClassTeam.class_id == class_id
    ).first()
    if not target_team:
        raise HTTPException(status_code=404, detail="目标队伍不存在")

    db.query(models.ClassTeamMember).filter(models.ClassTeamMember.student_id == data.student_id).delete()
    db.add(models.ClassTeamMember(team_id=target_team.id, student_id=data.student_id))
    db.commit()

    return {"message": "成员队伍已调整"}


@app.delete("/class-teams/{team_id}")
def delete_team(team_id: int, db: Session = Depends(get_db)):
    team = db.query(models.ClassTeam).filter(models.ClassTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")

    db.query(models.ClassTeamMember).filter(models.ClassTeamMember.team_id == team_id).delete()
    db.delete(team)
    db.commit()
    return {"message": "队伍已删除"}


# ================= 管理员中心：学期/老师/学生管理 =================
@app.post("/admin/semesters")
def create_semester(payload: schemas.SemesterCreate, db: Session = Depends(get_db)):
    if db.query(models.Semester).filter(models.Semester.name == payload.name).first():
        raise HTTPException(status_code=400, detail="学期名称已存在")

    if payload.is_current:
        db.query(models.Semester).update({models.Semester.is_current: False})

    item = models.Semester(
        name=payload.name,
        start_date=payload.start_date,
        end_date=payload.end_date,
        is_enabled=payload.is_enabled,
        is_current=payload.is_current,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"message": "学期创建成功", "id": item.id}


@app.put("/admin/semesters/{semester_id}")
def update_semester(semester_id: int, payload: schemas.SemesterUpdate, db: Session = Depends(get_db)):
    item = db.query(models.Semester).filter(models.Semester.id == semester_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="学期不存在")

    if payload.name is not None:
        dup = db.query(models.Semester).filter(models.Semester.name == payload.name, models.Semester.id != semester_id).first()
        if dup:
            raise HTTPException(status_code=400, detail="学期名称已存在")
        item.name = payload.name

    if payload.start_date is not None:
        item.start_date = payload.start_date
    if payload.end_date is not None:
        item.end_date = payload.end_date
    if payload.is_enabled is not None:
        item.is_enabled = payload.is_enabled

    if payload.is_current is not None:
        if payload.is_current:
            db.query(models.Semester).update({models.Semester.is_current: False})
            item.is_current = True
        else:
            item.is_current = False

    db.commit()
    return {"message": "学期更新成功"}


@app.delete("/admin/semesters/{semester_id}")
def delete_semester(semester_id: int, db: Session = Depends(get_db)):
    item = db.query(models.Semester).filter(models.Semester.id == semester_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="学期不存在")

    class_count = db.query(models.TeachingClass).filter(models.TeachingClass.semester_id == semester_id).count()
    if class_count > 0:
        raise HTTPException(status_code=400, detail="该学期下已有班级，无法删除")

    db.delete(item)
    db.commit()
    return {"message": "学期已删除"}


@app.get("/admin/teachers")
def admin_list_teachers(db: Session = Depends(get_db)):
    items = _list_distinct_teacher_users(db)
    return [
        {
            "id": u.id,
            "username": u.username,
            "real_name": u.real_name,
            "is_admin": (u.role == "admin") or bool(getattr(u, "is_admin", False)),
            "role": "admin" if ((u.role == "admin") or bool(getattr(u, "is_admin", False))) else "teacher",
        }
        for u in items
    ]


@app.post("/admin/teachers")
def admin_create_teacher(payload: schemas.TeacherCreate, db: Session = Depends(get_db)):
    normalized_username = _normalize_username(payload.username)
    if db.query(models.User).filter(models.User.username == normalized_username).first():
        raise HTTPException(status_code=400, detail="账号已存在")

    role = "admin" if payload.is_admin else "teacher"
    item = models.User(
        username=normalized_username,
        password=pwd_context.hash(payload.password),
        role=role,
        is_admin=payload.is_admin,
        real_name=payload.real_name,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"message": "老师账号创建成功", "id": item.id}


@app.put("/admin/teachers/{teacher_id}")
def admin_update_teacher(teacher_id: int, payload: schemas.TeacherUpdate, db: Session = Depends(get_db)):
    item = db.query(models.User).filter(models.User.id == teacher_id, models.User.role.in_(["teacher", "admin"])).first()
    if not item:
        raise HTTPException(status_code=404, detail="老师账号不存在")

    if payload.username is not None:
        normalized_username = _normalize_username(payload.username)
        if normalized_username != item.username:
            dup = db.query(models.User).filter(models.User.username == normalized_username, models.User.id != teacher_id).first()
            if dup:
                raise HTTPException(status_code=400, detail="账号已存在")
            item.username = normalized_username

    if payload.real_name is not None:
        item.real_name = payload.real_name

    if payload.password:
        item.password = pwd_context.hash(payload.password)

    if payload.is_admin is not None:
        item.is_admin = payload.is_admin
        item.role = "admin" if payload.is_admin else "teacher"

    db.commit()
    return {"message": "老师账号更新成功"}


@app.delete("/admin/teachers/{teacher_id}")
def admin_delete_teacher(teacher_id: int, db: Session = Depends(get_db)):
    item = db.query(models.User).filter(models.User.id == teacher_id, models.User.role.in_(["teacher", "admin"])).first()
    if not item:
        raise HTTPException(status_code=404, detail="老师账号不存在")

    class_count = db.query(models.TeachingClass).filter(models.TeachingClass.teacher_id == teacher_id).count()
    if class_count > 0:
        raise HTTPException(status_code=400, detail="该老师已关联班级，无法删除")

    db.delete(item)
    db.commit()
    return {"message": "老师账号已删除"}


@app.get("/admin/students")
def admin_list_students(
    grade: Optional[str] = Query(default=None),
    class_name: Optional[str] = Query(default=None),
    username: Optional[str] = Query(default=None),
    student_no: Optional[str] = Query(default=None),
    real_name: Optional[str] = Query(default=None),
    teacher_id: Optional[int] = Query(default=None),
    teacher_name: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    q = db.query(models.User).filter(models.User.role == "student")

    if grade:
        q = q.filter(models.User.grade.ilike(f"%{grade}%"))
    if username:
        q = q.filter(models.User.username.ilike(f"%{username}%"))
    if student_no:
        q = q.filter(models.User.student_no.ilike(f"%{student_no}%"))
    if real_name:
        q = q.filter(models.User.real_name.ilike(f"%{real_name}%"))

    students = q.order_by(models.User.id.desc()).all()
    student_ids = [s.id for s in students]
    current = _get_current_semester(db)

    teacher_ids = {s.teacher_id for s in students if s.teacher_id}
    relation_rows = []
    semester_name = current.name if current else None

    if current and student_ids:
        relation_rows = db.query(
            models.ClassStudent.student_id,
            models.ClassStudent.class_id,
            models.TeachingClass.name,
            models.TeachingClass.teacher_id
        ).join(
            models.TeachingClass, models.ClassStudent.class_id == models.TeachingClass.id
        ).filter(
            models.ClassStudent.student_id.in_(student_ids),
            models.TeachingClass.semester_id == current.id
        ).order_by(models.ClassStudent.id.desc()).all()
        teacher_ids.update(r.teacher_id for r in relation_rows if r.teacher_id)

    teacher_list = db.query(models.User).filter(models.User.id.in_(list(teacher_ids))).all() if teacher_ids else []
    teacher_map = {t.id: t for t in teacher_list}

    keyword_class = class_name.lower() if class_name else None
    keyword_teacher = teacher_name.lower() if teacher_name else None

    result = []
    relation_group: dict[int, list] = {}
    for r in relation_rows:
        relation_group.setdefault(r.student_id, []).append(r)

    for stu in students:
        rows = relation_group.get(stu.id, [])
        if rows:
            for rel in rows:
                teacher = teacher_map.get(rel.teacher_id) if rel.teacher_id else None
                final_teacher_name = (teacher.real_name or teacher.username) if teacher else None

                if teacher_id and rel.teacher_id != teacher_id:
                    continue
                if keyword_class and (not rel.name or keyword_class not in rel.name.lower()):
                    continue
                if keyword_teacher and (not final_teacher_name or keyword_teacher not in final_teacher_name.lower()):
                    continue

                result.append({
                    "id": stu.id,
                    "class_id": rel.class_id,
                    "semester": semester_name,
                    "username": stu.username,
                    "student_no": stu.student_no,
                    "real_name": stu.real_name,
                    "grade": stu.grade,
                    "class_name": rel.name,
                    "teacher_id": rel.teacher_id,
                    "teacher_name": final_teacher_name,
                })
            continue

        fallback_teacher = teacher_map.get(stu.teacher_id) if stu.teacher_id else None
        fallback_teacher_name = (fallback_teacher.real_name or fallback_teacher.username) if fallback_teacher else None

        if teacher_id and stu.teacher_id != teacher_id:
            continue
        if keyword_class and (not stu.class_name or keyword_class not in stu.class_name.lower()):
            continue
        if keyword_teacher and (not fallback_teacher_name or keyword_teacher not in fallback_teacher_name.lower()):
            continue

        result.append({
            "id": stu.id,
            "class_id": None,
            "semester": semester_name,
            "username": stu.username,
            "student_no": stu.student_no,
            "real_name": stu.real_name,
            "grade": stu.grade,
            "class_name": stu.class_name,
            "teacher_id": stu.teacher_id,
            "teacher_name": fallback_teacher_name,
        })

    return result



@app.post("/admin/students")
def admin_create_student(payload: schemas.StudentCreate, db: Session = Depends(get_db)):
    normalized_username = _normalize_username(payload.username)
    if db.query(models.User).filter(models.User.username == normalized_username).first():
        raise HTTPException(status_code=400, detail="账号已存在")
    if payload.student_no and db.query(models.User).filter(models.User.student_no == payload.student_no).first():
        raise HTTPException(status_code=400, detail="学号已存在")

    canonical_teacher_id = _canonical_teacher_id(db, payload.teacher_id)
    if payload.teacher_id is not None and canonical_teacher_id is None:
        raise HTTPException(status_code=400, detail="归属老师不存在")

    item = models.User(
        username=normalized_username,
        password=pwd_context.hash(payload.password),
        role="student",
        real_name=payload.real_name,
        student_no=payload.student_no,
        grade=payload.grade,
        class_name=payload.class_name,
        teacher_id=canonical_teacher_id,
    )
    try:
        db.add(item)
        db.flush()
        _ensure_student_class_relation(db, item)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="账号或学号已存在")

    db.refresh(item)
    return {"message": "学生账号创建成功", "id": item.id}




@app.put("/admin/students/{student_id}")
def admin_update_student(student_id: int, payload: schemas.StudentUpdate, db: Session = Depends(get_db)):
    item = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not item:
        raise HTTPException(status_code=404, detail="学生不存在")

    if payload.username is not None:
        normalized_username = _normalize_username(payload.username)
        if normalized_username != item.username:
            dup = db.query(models.User).filter(models.User.username == normalized_username, models.User.id != student_id).first()
            if dup:
                raise HTTPException(status_code=400, detail="账号已存在")
            item.username = normalized_username

    if payload.student_no is not None and payload.student_no != item.student_no:
        dup_no = db.query(models.User).filter(models.User.student_no == payload.student_no, models.User.id != student_id).first()
        if dup_no:
            raise HTTPException(status_code=400, detail="学号已存在")
        item.student_no = payload.student_no

    if payload.real_name is not None:
        item.real_name = payload.real_name
    if payload.grade is not None:
        item.grade = payload.grade
    if payload.class_name is not None:
        item.class_name = payload.class_name
    if payload.teacher_id is not None:
        canonical_teacher_id = _canonical_teacher_id(db, payload.teacher_id)
        if canonical_teacher_id is None:
            raise HTTPException(status_code=400, detail="归属老师不存在")
        item.teacher_id = canonical_teacher_id
    if payload.password:
        item.password = pwd_context.hash(payload.password)

    try:
        _ensure_student_class_relation(db, item)
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="账号或学号已存在")
    return {"message": "学生信息更新成功"}




@app.delete("/admin/students/{student_id}")
def admin_delete_student(student_id: int, db: Session = Depends(get_db)):
    item = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not item:
        raise HTTPException(status_code=404, detail="学生不存在")

    db.query(models.ClassStudent).filter(models.ClassStudent.student_id == student_id).delete()
    db.query(models.ClassTeamMember).filter(models.ClassTeamMember.student_id == student_id).delete()
    db.delete(item)
    db.commit()
    return {"message": "学生账号已删除"}


@app.put("/admin/students/{student_id}/password")
def admin_reset_student_password(student_id: int, payload: schemas.PasswordReset, db: Session = Depends(get_db)):
    item = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not item:
        raise HTTPException(status_code=404, detail="学生不存在")

    item.password = pwd_context.hash(payload.new_password)
    db.commit()
    return {"message": "密码已重置"}


@app.get("/teachers/{teacher_id}/students")
def teacher_student_list(
    teacher_id: int,
    class_name: Optional[str] = Query(default=None),
    teacher_name: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    teacher = _require_teacher_user(db, teacher_id)
    teacher_id = teacher.id

    q = db.query(models.User).filter(models.User.role == "student", models.User.teacher_id == teacher_id)
    if class_name:
        q = q.filter(models.User.class_name.ilike(f"%{class_name}%"))
    if teacher_name:
        display_name = teacher.real_name or teacher.username
        if teacher_name.lower() not in display_name.lower():
            return []

    items = q.order_by(models.User.id.desc()).all()
    return [
        {
            "id": u.id,
            "username": u.username,
            "student_no": u.student_no,
            "real_name": u.real_name,
            "grade": u.grade,
            "class_name": u.class_name,
            "teacher_id": teacher_id,
            "teacher_name": teacher.real_name or teacher.username,
        }
        for u in items
    ]


def _ensure_openpyxl():
    if load_workbook is None or Workbook is None:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请先安装：pip install openpyxl")


@app.get("/admin/templates/teachers")
def teacher_template():
    _ensure_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "teachers"
    ws.append(["username", "password", "real_name", "is_admin"])
    ws.append(["teacher001", "123456", "张老师", False])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=teachers_template.xlsx"},
    )


@app.get("/admin/templates/students")
def student_template():
    _ensure_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(["username", "password", "student_no", "real_name", "grade", "class_name", "teacher_username"])
    ws.append(["stu001", "123456", "20260001", "李同学", "2026", "软件1班", "teacher001"])


    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_template.xlsx"},
    )


@app.get("/admin/templates/classes")
def class_template():
    _ensure_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "classes"
    ws.append(["semester_id", "semester_name", "class_name", "teacher_username", "teacher_real_name"])
    ws.append([1, "", "软件1班", "teacher001", ""])
    ws.append(["", "2026-秋", "软件2班", "", "杨珊"])

    note_ws = wb.create_sheet("说明")
    note_ws.append(["字段", "说明"])
    note_ws.append(["semester_id", "学期ID，优先使用；填了则按ID匹配学期"])
    note_ws.append(["semester_name", "学期名称；semester_id 为空时使用（如：2026-秋）"])
    note_ws.append(["class_name", "班级名称，必填"])
    note_ws.append(["teacher_username", "教师账号，优先使用"])
    note_ws.append(["teacher_real_name", "教师姓名；teacher_username 为空时兜底匹配（需唯一）"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=classes_template.xlsx"},
    )


@app.post("/admin/teachers/import")
async def import_teachers(file: UploadFile = File(...), db: Session = Depends(get_db)):

    _ensure_openpyxl()
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        username, password, real_name, is_admin = row[:4]
        username = str(username or "").strip()
        password = str(password or "").strip()
        real_name = str(real_name or "").strip() or None

        if not username or not password:
            skipped += 1
            continue

        if db.query(models.User).filter(models.User.username == username).first():
            skipped += 1
            continue

        admin_flag = str(is_admin).strip().lower() in ["1", "true", "yes", "是"] if is_admin is not None else False
        db.add(models.User(
            username=username,
            password=pwd_context.hash(password),
            role="admin" if admin_flag else "teacher",
            is_admin=admin_flag,
            real_name=real_name,
        ))
        created += 1

    db.commit()
    return {"message": "导入完成", "created": created, "skipped": skipped}


@app.post("/admin/students/import")
async def import_students(file: UploadFile = File(...), db: Session = Depends(get_db)):
    _ensure_openpyxl()
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        username, password, student_no, real_name, grade, class_name, teacher_username = row[:7]
        username = str(username or "").strip()
        password = str(password or "").strip()
        student_no = str(student_no or "").strip() or None
        real_name = str(real_name or "").strip() or None
        grade = str(grade or "").strip() or None
        class_name = str(class_name or "").strip() or None
        teacher_username = str(teacher_username or "").strip() or None


        if not username or not password:
            skipped += 1
            continue

        if db.query(models.User).filter(models.User.username == username).first():
            skipped += 1
            continue

        if student_no and db.query(models.User).filter(models.User.student_no == student_no).first():
            skipped += 1
            continue

        teacher_id = None
        if teacher_username:
            teacher = db.query(models.User).filter(models.User.username == teacher_username, models.User.role.in_(["teacher", "admin"])).first()
            if not teacher:
                skipped += 1
                continue
            teacher_id = teacher.id

        db.add(models.User(
            username=username,
            password=pwd_context.hash(password),
            role="student",
            real_name=real_name,
            student_no=student_no,
            grade=grade,
            class_name=class_name,
            teacher_id=teacher_id,
        ))
        created += 1

    db.commit()
    return {"message": "导入完成", "created": created, "skipped": skipped}


@app.post("/admin/classes/import")
async def import_classes(file: UploadFile = File(...), db: Session = Depends(get_db)):
    _ensure_openpyxl()
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="模板为空")

    headers = [str(x or "").strip() for x in header_row]
    col_index = {name: idx for idx, name in enumerate(headers) if name}

    def _value(row_values, key: str) -> str:
        idx = col_index.get(key)
        if idx is None or idx >= len(row_values):
            return ""
        return str(row_values[idx] or "").strip()

    created = 0
    skipped = 0
    for _, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_values = list(row)

        semester_id_text = _value(row_values, "semester_id")
        semester_name = _value(row_values, "semester_name")
        class_name = _value(row_values, "class_name")
        teacher_username = _value(row_values, "teacher_username")
        teacher_real_name = _value(row_values, "teacher_real_name")

        # 兼容旧模板（三列：semester_name, class_name, teacher_username）
        if not class_name and len(headers) >= 3 and headers[:3] == ["semester_name", "class_name", "teacher_username"]:
            semester_name = str(row_values[0] or "").strip()
            class_name = str(row_values[1] or "").strip()
            teacher_username = str(row_values[2] or "").strip()

        if not class_name:
            skipped += 1
            continue

        semester = None
        if semester_id_text:
            try:
                semester_id = int(float(semester_id_text))
            except ValueError:
                semester_id = None
            if semester_id is not None:
                semester = db.query(models.Semester).filter(models.Semester.id == semester_id).first()

        if not semester and semester_name:
            semester = db.query(models.Semester).filter(models.Semester.name == semester_name).first()

        if not semester:
            skipped += 1
            continue

        teacher_user = None
        if teacher_username:
            teacher_user = _find_user_by_username(db, teacher_username)

        if not teacher_user and teacher_real_name:
            teacher_candidates = db.query(models.User).filter(
                models.User.role.in_(["teacher", "admin"]),
                models.User.real_name == teacher_real_name
            ).all()
            if len(teacher_candidates) == 1:
                teacher_user = teacher_candidates[0]

        if not teacher_user or teacher_user.role not in ["teacher", "admin"]:
            skipped += 1
            continue

        canonical_teacher_id = _canonical_teacher_id(db, teacher_user.id)
        if canonical_teacher_id is None:
            skipped += 1
            continue

        existed = db.query(models.TeachingClass).filter(
            models.TeachingClass.semester_id == semester.id,
            models.TeachingClass.name == class_name,
            models.TeachingClass.teacher_id == canonical_teacher_id
        ).first()
        if existed:
            skipped += 1
            continue

        db.add(models.TeachingClass(
            name=class_name,
            semester_id=semester.id,
            teacher_id=canonical_teacher_id
        ))
        created += 1

    db.commit()
    current = _get_current_semester(db)
    if current:
        _sync_teacher_classes_from_students(db, current.id)
    return {"message": "导入完成", "created": created, "skipped": skipped}


# ================= 综合项目组队申请体系 =================
def _parse_date_only(text: Optional[str]):
    value = (text or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def _get_or_create_project_setting(db: Session, semester_id: int):
    item = db.query(models.ProjectApplicationSetting).filter(
        models.ProjectApplicationSetting.semester_id == semester_id
    ).first()
    if item:
        return item

    item = models.ProjectApplicationSetting(
        semester_id=semester_id,
        is_enabled=False,
        min_team_size=2,
        max_team_size=6,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def _is_project_apply_open(setting: models.ProjectApplicationSetting):
    if not setting or not setting.is_enabled:
        return False

    today = datetime.now().date()
    start = _parse_date_only(setting.open_start)
    end = _parse_date_only(setting.open_end)
    if start and today < start:
        return False
    if end and today > end:
        return False
    return True


def _generate_project_team_no(db: Session, semester_id: int, class_id: int):
    existed = db.query(models.ProjectTeam).filter(
        models.ProjectTeam.semester_id == semester_id,
        models.ProjectTeam.class_id == class_id
    ).count()
    return f"C{class_id:03d}-T{existed + 1:03d}"


def _get_project_team_member_users(db: Session, team_id: int):
    links = db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.team_id == team_id).all()
    student_ids = [x.student_id for x in links]
    users = db.query(models.User).filter(models.User.id.in_(student_ids)).all() if student_ids else []
    user_map = {u.id: u for u in users}

    rows = []
    for sid in student_ids:
        u = user_map.get(sid)
        if not u:
            continue
        rows.append({
            "id": u.id,
            "username": u.username,
            "real_name": u.real_name,
            "student_no": u.student_no,
            "class_name": u.class_name,
        })
    return rows


def _project_team_to_dict(db: Session, team: models.ProjectTeam):
    topic = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == team.topic_id).first() if team.topic_id else None
    teacher = db.query(models.User).filter(models.User.id == team.advisor_teacher_id).first() if team.advisor_teacher_id else None
    cls = db.query(models.TeachingClass).filter(models.TeachingClass.id == team.class_id).first()
    semester = db.query(models.Semester).filter(models.Semester.id == team.semester_id).first()

    return {
        "id": team.id,
        "team_no": team.team_no,
        "team_name": team.team_name,
        "direction": team.direction,
        "status": team.status,
        "review_comment": team.review_comment,
        "class_id": team.class_id,
        "class_name": cls.name if cls else None,
        "semester_id": team.semester_id,
        "semester_name": semester.name if semester else None,
        "leader_id": team.leader_id,
        "topic_id": team.topic_id,
        "topic_name": topic.title if topic else None,
        "advisor_teacher_id": team.advisor_teacher_id,
        "advisor_teacher_name": (teacher.real_name or teacher.username) if teacher else None,
        "apply_note": team.apply_note,
        "apply_attachment_path": team.apply_attachment_path,
        "created_at": team.created_at.isoformat() if team.created_at else None,
        "members": _get_project_team_member_users(db, team.id)
    }


def _ensure_student_single_team_in_semester(db: Session, student_id: int, semester_id: int, ignore_team_id: Optional[int] = None):
    q = db.query(models.ProjectTeamMember, models.ProjectTeam).join(
        models.ProjectTeam, models.ProjectTeamMember.team_id == models.ProjectTeam.id
    ).filter(
        models.ProjectTeamMember.student_id == student_id,
        models.ProjectTeam.semester_id == semester_id
    )
    if ignore_team_id:
        q = q.filter(models.ProjectTeam.id != ignore_team_id)
    existed = q.first()
    if existed:
        raise HTTPException(status_code=400, detail="该学生在当前学期已加入其他队伍")


@app.post("/project/upload")
def upload_project_file(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    allowed_ext = {".pdf", ".doc", ".docx", ".zip", ".rar", ".7z", ".txt", ".md", ".ppt", ".pptx"}
    if ext and ext not in allowed_ext:
        raise HTTPException(status_code=400, detail="不支持的附件类型")

    safe_name = _safe_file_part(Path(file.filename or "file").stem, "project_file")
    folder = Path(__file__).resolve().parent / "document" / "project_uploads"
    folder.mkdir(parents=True, exist_ok=True)
    final_path = folder / f"{int(time.time() * 1000)}_{safe_name}{ext or '.dat'}"
    with final_path.open("wb") as f:
        f.write(file.file.read())

    return {"path": final_path.as_posix(), "filename": file.filename}


@app.get("/file/preview")
def preview_local_file(path: str = Query(...)):
    file_path = Path(path)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(file_path)


@app.get("/admin/project/topics")

def admin_list_project_topics(
    semester_id: Optional[int] = Query(default=None),
    keyword: Optional[str] = Query(default=None),
    direction: Optional[str] = Query(default=None),
    teacher_id: Optional[int] = Query(default=None),
    only_current: bool = Query(default=True),
    db: Session = Depends(get_db)
):
    current = _get_current_semester(db)
    sid = semester_id
    if only_current and current:
        sid = current.id

    q = db.query(models.ProjectTopic)
    if sid:
        q = q.filter(models.ProjectTopic.semester_id == sid)
    if teacher_id:
        q = q.filter(models.ProjectTopic.teacher_id == teacher_id)
    if keyword:
        q = q.filter(models.ProjectTopic.title.ilike(f"%{keyword}%"))
    if direction:
        q = q.filter(models.ProjectTopic.direction.ilike(f"%{direction}%"))

    rows = q.order_by(models.ProjectTopic.id.desc()).all()
    teacher_ids = list({x.teacher_id for x in rows})
    teachers = db.query(models.User).filter(models.User.id.in_(teacher_ids)).all() if teacher_ids else []
    teacher_map = {t.id: (t.real_name or t.username) for t in teachers}

    return [
        {
            "id": x.id,
            "semester_id": x.semester_id,
            "title": x.title,
            "teacher_id": x.teacher_id,
            "teacher_name": teacher_map.get(x.teacher_id),
            "materials": x.materials,
            "attachment_path": x.attachment_path,
            "direction": x.direction,
            "is_published": x.is_published,
            "created_by": x.created_by,
            "created_at": x.created_at.isoformat() if x.created_at else None,
        }
        for x in rows
    ]


@app.post("/admin/project/topics")
def admin_create_project_topic(payload: schemas.ProjectTopicCreate, db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    teacher = _require_teacher_user(db, payload.teacher_id)

    item = models.ProjectTopic(
        semester_id=current.id,
        title=payload.title.strip(),
        teacher_id=teacher.id,
        materials=(payload.materials or "").strip() or None,
        attachment_path=(payload.attachment_path or "").strip() or None,
        direction=payload.direction.strip(),
        is_published=payload.is_published,
        created_by="admin"
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"message": "课题创建成功", "id": item.id}


@app.put("/admin/project/topics/{topic_id}")
def admin_update_project_topic(topic_id: int, payload: schemas.ProjectTopicUpdate, db: Session = Depends(get_db)):
    item = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == topic_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="课题不存在")

    if payload.title is not None:
        item.title = payload.title.strip()
    if payload.teacher_id is not None:
        teacher = _require_teacher_user(db, payload.teacher_id)
        item.teacher_id = teacher.id
    if payload.materials is not None:
        item.materials = (payload.materials or "").strip() or None
    if payload.attachment_path is not None:
        item.attachment_path = (payload.attachment_path or "").strip() or None
    if payload.direction is not None:
        item.direction = payload.direction.strip()
    if payload.is_published is not None:
        item.is_published = payload.is_published

    db.commit()
    return {"message": "课题更新成功"}


@app.delete("/admin/project/topics/{topic_id}")
def admin_delete_project_topic(topic_id: int, db: Session = Depends(get_db)):
    item = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == topic_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="课题不存在")

    used = db.query(models.ProjectTeam).filter(models.ProjectTeam.topic_id == topic_id).count()
    if used > 0:
        raise HTTPException(status_code=400, detail="该课题已被队伍申请，无法删除")

    db.query(models.ProjectClassTopicConfig).filter(models.ProjectClassTopicConfig.topic_id == topic_id).delete(synchronize_session=False)
    db.delete(item)
    db.commit()
    return {"message": "课题已删除"}


@app.get("/admin/project/settings/current")
def admin_get_project_setting(db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    item = _get_or_create_project_setting(db, current.id)
    return {
        "semester_id": item.semester_id,
        "is_enabled": item.is_enabled,
        "open_start": item.open_start,
        "open_end": item.open_end,
        "min_team_size": item.min_team_size,
        "max_team_size": item.max_team_size,
    }


@app.put("/admin/project/settings/current")
def admin_update_project_setting(payload: schemas.ProjectApplicationSettingUpdate, db: Session = Depends(get_db)):
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    if payload.min_team_size <= 0:
        raise HTTPException(status_code=400, detail="队伍人数下限必须大于0")
    if payload.max_team_size < payload.min_team_size:
        raise HTTPException(status_code=400, detail="队伍人数上限不能小于下限")

    item = _get_or_create_project_setting(db, current.id)
    item.is_enabled = payload.is_enabled
    item.open_start = (payload.open_start or "").strip() or None
    item.open_end = (payload.open_end or "").strip() or None
    item.min_team_size = payload.min_team_size
    item.max_team_size = payload.max_team_size
    db.commit()
    return {"message": "综设申请设置已保存"}


@app.get("/admin/project/class-topic-configs")
def admin_get_class_topic_configs(
    semester_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db)
):
    current = _get_current_semester(db)
    sid = semester_id or (current.id if current else None)
    if sid is None:
        return []

    classes = db.query(models.TeachingClass).filter(models.TeachingClass.semester_id == sid).order_by(models.TeachingClass.id.asc()).all()
    class_ids = [x.id for x in classes]
    cfg_rows = db.query(models.ProjectClassTopicConfig).filter(models.ProjectClassTopicConfig.class_id.in_(class_ids)).all() if class_ids else []
    topic_ids = [x.topic_id for x in cfg_rows]
    topics = db.query(models.ProjectTopic).filter(models.ProjectTopic.id.in_(topic_ids)).all() if topic_ids else []

    topic_map = {t.id: t for t in topics}
    class_cfg_map = {}
    for c in cfg_rows:
        class_cfg_map.setdefault(c.class_id, []).append(c.topic_id)

    return [
        {
            "class_id": cls.id,
            "class_name": cls.name,
            "topic_ids": class_cfg_map.get(cls.id, []),
            "topics": [
                {
                    "id": tid,
                    "title": topic_map[tid].title,
                    "direction": topic_map[tid].direction,
                    "teacher_id": topic_map[tid].teacher_id,
                    "is_published": topic_map[tid].is_published,
                }
                for tid in class_cfg_map.get(cls.id, []) if tid in topic_map
            ]
        }
        for cls in classes
    ]


@app.put("/admin/project/classes/{class_id}/topics")
def admin_set_class_topics(class_id: int, payload: schemas.ProjectClassTopicAssign, db: Session = Depends(get_db)):
    cls = db.query(models.TeachingClass).filter(models.TeachingClass.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    topic_ids = list(set(payload.topic_ids or []))
    if topic_ids:
        valid_count = db.query(models.ProjectTopic).filter(models.ProjectTopic.id.in_(topic_ids)).count()
        if valid_count != len(topic_ids):
            raise HTTPException(status_code=400, detail="包含无效课题")

    db.query(models.ProjectClassTopicConfig).filter(models.ProjectClassTopicConfig.class_id == class_id).delete(synchronize_session=False)
    for tid in topic_ids:
        db.add(models.ProjectClassTopicConfig(class_id=class_id, topic_id=tid))

    db.commit()
    return {"message": "班级课题配置已更新", "topic_count": len(topic_ids)}


@app.get("/admin/project/teams")
def admin_list_project_teams(
    semester_id: Optional[int] = Query(default=None),
    direction: Optional[str] = Query(default=None),
    topic_keyword: Optional[str] = Query(default=None),
    teacher_keyword: Optional[str] = Query(default=None),
    team_keyword: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    q = db.query(models.ProjectTeam)
    if semester_id:
        q = q.filter(models.ProjectTeam.semester_id == semester_id)
    if direction:
        q = q.filter(models.ProjectTeam.direction.ilike(f"%{direction}%"))
    if team_keyword:
        q = q.filter(
            (models.ProjectTeam.team_no.ilike(f"%{team_keyword}%")) |
            (models.ProjectTeam.team_name.ilike(f"%{team_keyword}%"))
        )

    teams = q.order_by(models.ProjectTeam.id.desc()).all()
    rows = [_project_team_to_dict(db, t) for t in teams]

    if topic_keyword:
        rows = [x for x in rows if x.get("topic_name") and topic_keyword.lower() in x["topic_name"].lower()]
    if teacher_keyword:
        rows = [x for x in rows if x.get("advisor_teacher_name") and teacher_keyword.lower() in x["advisor_teacher_name"].lower()]

    return rows


@app.post("/admin/project/teams")
def admin_create_project_team(payload: schemas.ProjectTeamAdminCreate, db: Session = Depends(get_db)):
    cls = db.query(models.TeachingClass).filter(models.TeachingClass.id == payload.class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    leader = db.query(models.User).filter(models.User.id == payload.leader_id, models.User.role == "student").first()
    if not leader:
        raise HTTPException(status_code=404, detail="队长不存在")

    _ensure_student_single_team_in_semester(db, leader.id, cls.semester_id)

    team = models.ProjectTeam(
        semester_id=cls.semester_id,
        class_id=cls.id,
        team_no=_generate_project_team_no(db, cls.semester_id, cls.id),
        team_name=payload.team_name.strip(),
        direction=payload.direction.strip(),
        leader_id=leader.id,
        status="draft",
    )
    db.add(team)
    db.flush()

    member_ids = list(dict.fromkeys([leader.id] + [x for x in (payload.member_ids or []) if x != leader.id]))
    for sid in member_ids:
        stu = db.query(models.User).filter(models.User.id == sid, models.User.role == "student").first()
        if not stu:
            continue
        _ensure_student_single_team_in_semester(db, sid, cls.semester_id, ignore_team_id=team.id)
        db.add(models.ProjectTeamMember(team_id=team.id, student_id=sid))

    db.commit()
    return {"message": "队伍创建成功", "team_id": team.id}


@app.put("/admin/project/teams/{team_id}")
def admin_update_project_team(team_id: int, payload: schemas.ProjectTeamUpdate, db: Session = Depends(get_db)):
    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")

    if payload.team_name is not None:
        team.team_name = payload.team_name.strip()
    if payload.direction is not None:
        team.direction = payload.direction.strip()

    if payload.member_ids is not None:
        member_ids = list(dict.fromkeys(payload.member_ids))
        if team.leader_id not in member_ids:
            member_ids = [team.leader_id] + member_ids

        db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.team_id == team_id).delete(synchronize_session=False)
        for sid in member_ids:
            stu = db.query(models.User).filter(models.User.id == sid, models.User.role == "student").first()
            if not stu:
                continue
            _ensure_student_single_team_in_semester(db, sid, team.semester_id, ignore_team_id=team.id)
            db.add(models.ProjectTeamMember(team_id=team.id, student_id=sid))

    db.commit()
    return {"message": "队伍信息已更新"}


@app.delete("/admin/project/teams/{team_id}")
def admin_delete_project_team(team_id: int, db: Session = Depends(get_db)):
    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")

    db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.team_id == team_id).delete(synchronize_session=False)
    db.delete(team)
    db.commit()
    return {"message": "队伍已删除"}


@app.get("/teachers/{teacher_id}/project/topics")
def teacher_list_project_topics(
    teacher_id: int,
    semester_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id
    current = _get_current_semester(db)
    sid = semester_id or (current.id if current else None)

    q = db.query(models.ProjectTopic).filter(models.ProjectTopic.teacher_id == teacher_id)
    if sid:
        q = q.filter(models.ProjectTopic.semester_id == sid)

    rows = q.order_by(models.ProjectTopic.id.desc()).all()
    return [
        {
            "id": x.id,
            "semester_id": x.semester_id,
            "title": x.title,
            "teacher_id": x.teacher_id,
            "materials": x.materials,
            "attachment_path": x.attachment_path,
            "direction": x.direction,
            "is_published": x.is_published,
            "created_by": x.created_by,
            "created_at": x.created_at.isoformat() if x.created_at else None,
        }
        for x in rows
    ]


@app.post("/teachers/{teacher_id}/project/topics")
def teacher_create_project_topic(teacher_id: int, payload: schemas.ProjectTopicCreate, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id
    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    item = models.ProjectTopic(
        semester_id=current.id,
        title=payload.title.strip(),
        teacher_id=teacher_id,
        materials=(payload.materials or "").strip() or None,
        attachment_path=(payload.attachment_path or "").strip() or None,
        direction=payload.direction.strip(),
        is_published=payload.is_published,
        created_by="teacher"
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"message": "题目创建成功", "id": item.id}


@app.put("/teachers/{teacher_id}/project/topics/{topic_id}")
def teacher_update_project_topic(teacher_id: int, topic_id: int, payload: schemas.ProjectTopicUpdate, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id
    item = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == topic_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="课题不存在")
    if item.teacher_id != teacher_id:
        raise HTTPException(status_code=403, detail="无权修改该课题")

    if payload.title is not None:
        item.title = payload.title.strip()
    if payload.materials is not None:
        item.materials = (payload.materials or "").strip() or None
    if payload.attachment_path is not None:
        item.attachment_path = (payload.attachment_path or "").strip() or None
    if payload.direction is not None:
        item.direction = payload.direction.strip()
    if payload.is_published is not None:
        item.is_published = payload.is_published

    db.commit()
    return {"message": "题目更新成功"}


@app.delete("/teachers/{teacher_id}/project/topics/{topic_id}")
def teacher_delete_project_topic(teacher_id: int, topic_id: int, db: Session = Depends(get_db)):
    teacher_id = _require_teacher_user(db, teacher_id).id
    item = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == topic_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="课题不存在")
    if item.teacher_id != teacher_id:
        raise HTTPException(status_code=403, detail="无权删除该课题")

    used = db.query(models.ProjectTeam).filter(models.ProjectTeam.topic_id == topic_id).count()
    if used > 0:
        raise HTTPException(status_code=400, detail="该课题已有申请，无法删除")

    db.query(models.ProjectClassTopicConfig).filter(models.ProjectClassTopicConfig.topic_id == topic_id).delete(synchronize_session=False)
    db.delete(item)
    db.commit()
    return {"message": "题目已删除"}


@app.get("/teachers/{teacher_id}/project/applications")
def teacher_list_project_applications(
    teacher_id: int,
    status: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id
    current = _get_current_semester(db)
    if not current:
        return []

    q = db.query(models.ProjectTeam).filter(
        models.ProjectTeam.semester_id == current.id,
        models.ProjectTeam.advisor_teacher_id == teacher_id,
        models.ProjectTeam.topic_id.isnot(None)
    )
    if status:
        q = q.filter(models.ProjectTeam.status == status)

    teams = q.order_by(models.ProjectTeam.id.desc()).all()
    return [_project_team_to_dict(db, t) for t in teams]


@app.post("/teachers/{teacher_id}/project/applications/{team_id}/accept")
def teacher_accept_project_application(
    teacher_id: int,
    team_id: int,
    payload: schemas.ProjectTeamReviewPayload,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id
    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="申请队伍不存在")
    if team.advisor_teacher_id != teacher_id:
        raise HTTPException(status_code=403, detail="无权处理该申请")

    team.status = "accepted"
    team.review_comment = (payload.comment or "").strip() or None
    db.commit()
    return {"message": "已接受组队申请"}


@app.post("/teachers/{teacher_id}/project/applications/{team_id}/reject")
def teacher_reject_project_application(
    teacher_id: int,
    team_id: int,
    payload: schemas.ProjectTeamReviewPayload,
    db: Session = Depends(get_db)
):
    teacher_id = _require_teacher_user(db, teacher_id).id
    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="申请队伍不存在")
    if team.advisor_teacher_id != teacher_id:
        raise HTTPException(status_code=403, detail="无权处理该申请")

    team.status = "rejected"
    team.review_comment = (payload.comment or "").strip() or None
    db.commit()
    return {"message": "已拒绝组队申请"}


@app.get("/students/{student_id}/project/entry")
def student_project_entry_info(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        return {"setting": None, "classes": []}

    setting = _get_or_create_project_setting(db, current.id)
    links = db.query(models.ClassStudent).filter(models.ClassStudent.student_id == student_id).all()
    classes = db.query(models.TeachingClass).filter(
        models.TeachingClass.id.in_([x.class_id for x in links]) if links else False,
        models.TeachingClass.semester_id == current.id
    ).all() if links else []

    return {
        "setting": {
            "is_enabled": setting.is_enabled,
            "open_start": setting.open_start,
            "open_end": setting.open_end,
            "min_team_size": setting.min_team_size,
            "max_team_size": setting.max_team_size,
            "is_window_open": _is_project_apply_open(setting),
        },
        "classes": [
            {
                "id": c.id,
                "name": c.name,
                "teacher_id": c.teacher_id,
            }
            for c in classes
        ]
    }


@app.get("/students/{student_id}/project/students/search")
def student_search_project_students(
    student_id: int,
    class_id: int = Query(...),
    keyword: str = Query(...),
    db: Session = Depends(get_db)
):
    _ = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not _:
        raise HTTPException(status_code=404, detail="学生不存在")

    links = db.query(models.ClassStudent).filter(models.ClassStudent.class_id == class_id).all()
    student_ids = [x.student_id for x in links]
    if not student_ids:
        return []

    key = keyword.strip()
    if not key:
        return []

    rows = db.query(models.User).filter(
        models.User.id.in_(student_ids),
        models.User.role == "student",
        (
            models.User.student_no.ilike(f"%{key}%") |
            models.User.real_name.ilike(f"%{key}%") |
            models.User.username.ilike(f"%{key}%")
        )
    ).order_by(models.User.id.asc()).all()

    return [
        {
            "id": s.id,
            "username": s.username,
            "real_name": s.real_name,
            "student_no": s.student_no,
            "class_name": s.class_name,
        }
        for s in rows
    ]


@app.get("/students/{student_id}/project/teams")
def student_list_project_teams(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        return []

    my_team_links = db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.student_id == student_id).all()
    team_ids = [x.team_id for x in my_team_links]
    teams = db.query(models.ProjectTeam).filter(
        models.ProjectTeam.id.in_(team_ids) if team_ids else False,
        models.ProjectTeam.semester_id == current.id
    ).order_by(models.ProjectTeam.id.desc()).all() if team_ids else []

    return [_project_team_to_dict(db, t) for t in teams]


@app.post("/students/{student_id}/project/teams")
def student_create_project_team(student_id: int, payload: schemas.ProjectTeamCreate, db: Session = Depends(get_db)):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    current = _get_current_semester(db)
    if not current:
        raise HTTPException(status_code=400, detail="当前学期未发布")

    setting = _get_or_create_project_setting(db, current.id)
    if not _is_project_apply_open(setting):
        raise HTTPException(status_code=400, detail="当前不在综设组队开放时间")

    cls = db.query(models.TeachingClass).filter(
        models.TeachingClass.id == payload.class_id,
        models.TeachingClass.semester_id == current.id
    ).first()
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在或不在当前学期")

    link = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == cls.id,
        models.ClassStudent.student_id == student_id
    ).first()
    if not link:
        raise HTTPException(status_code=403, detail="你不在该班级，不能创建队伍")

    _ensure_student_single_team_in_semester(db, student_id, current.id)

    direction = payload.direction.strip()
    if student.class_name and direction != student.class_name:
        raise HTTPException(status_code=400, detail="队伍方向需与队长专业方向一致")

    team = models.ProjectTeam(
        semester_id=current.id,
        class_id=cls.id,
        team_no=_generate_project_team_no(db, current.id, cls.id),
        team_name=payload.team_name.strip(),
        direction=direction,
        leader_id=student_id,
        status="draft"
    )
    db.add(team)
    db.flush()
    db.add(models.ProjectTeamMember(team_id=team.id, student_id=student_id))
    db.commit()
    db.refresh(team)
    return {"message": "队伍创建成功", "team_id": team.id}


@app.post("/students/{student_id}/project/teams/{team_id}/members")
def student_add_project_team_member(
    student_id: int,
    team_id: int,
    payload: schemas.ProjectTeamMemberAdd,
    db: Session = Depends(get_db)
):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")
    if team.leader_id != student_id:
        raise HTTPException(status_code=403, detail="仅队长可添加成员")

    current = _get_current_semester(db)
    if not current or team.semester_id != current.id:
        raise HTTPException(status_code=400, detail="仅支持当前学期队伍操作")

    setting = _get_or_create_project_setting(db, current.id)
    if not _is_project_apply_open(setting):
        raise HTTPException(status_code=400, detail="当前不在综设组队开放时间")

    member_user = db.query(models.User).filter(models.User.id == payload.student_id, models.User.role == "student").first()
    if not member_user:
        raise HTTPException(status_code=404, detail="待添加学生不存在")

    class_link = db.query(models.ClassStudent).filter(
        models.ClassStudent.class_id == team.class_id,
        models.ClassStudent.student_id == payload.student_id
    ).first()
    if not class_link:
        raise HTTPException(status_code=400, detail="该学生不在同一班级")

    if (member_user.class_name or "") != team.direction:
        raise HTTPException(status_code=400, detail="该学生专业方向与队伍方向不一致")

    _ensure_student_single_team_in_semester(db, payload.student_id, team.semester_id)

    existed = db.query(models.ProjectTeamMember).filter(
        models.ProjectTeamMember.team_id == team_id,
        models.ProjectTeamMember.student_id == payload.student_id
    ).first()
    if existed:
        return {"message": "该成员已在队伍中"}

    current_count = db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.team_id == team_id).count()
    if current_count >= setting.max_team_size:
        raise HTTPException(status_code=400, detail=f"队伍人数不能超过{setting.max_team_size}人")

    db.add(models.ProjectTeamMember(team_id=team_id, student_id=payload.student_id))
    db.commit()
    return {"message": "成员添加成功"}


@app.delete("/students/{student_id}/project/teams/{team_id}/members/{member_id}")
def student_remove_project_team_member(
    student_id: int,
    team_id: int,
    member_id: int,
    db: Session = Depends(get_db)
):
    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")
    if team.leader_id != student_id:
        raise HTTPException(status_code=403, detail="仅队长可移除成员")
    if member_id == student_id:
        raise HTTPException(status_code=400, detail="队长不能移除自己")

    link = db.query(models.ProjectTeamMember).filter(
        models.ProjectTeamMember.team_id == team_id,
        models.ProjectTeamMember.student_id == member_id
    ).first()
    if not link:
        raise HTTPException(status_code=404, detail="成员不存在")

    db.delete(link)
    db.commit()
    return {"message": "成员已移除"}


@app.get("/students/{student_id}/project/topics/available")
def student_list_available_topics(
    student_id: int,
    team_id: int = Query(...),
    db: Session = Depends(get_db)
):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")

    me_in_team = db.query(models.ProjectTeamMember).filter(
        models.ProjectTeamMember.team_id == team_id,
        models.ProjectTeamMember.student_id == student_id
    ).first()
    if not me_in_team:
        raise HTTPException(status_code=403, detail="你不在该队伍中")

    cfg_rows = db.query(models.ProjectClassTopicConfig).filter(
        models.ProjectClassTopicConfig.class_id == team.class_id
    ).all()
    topic_ids = [x.topic_id for x in cfg_rows]
    if not topic_ids:
        return []

    topics = db.query(models.ProjectTopic).filter(
        models.ProjectTopic.id.in_(topic_ids),
        models.ProjectTopic.is_published == True,
        models.ProjectTopic.direction == team.direction
    ).order_by(models.ProjectTopic.id.desc()).all()

    teacher_ids = list({x.teacher_id for x in topics})
    teachers = db.query(models.User).filter(models.User.id.in_(teacher_ids)).all() if teacher_ids else []
    teacher_map = {t.id: (t.real_name or t.username) for t in teachers}

    return [
        {
            "id": x.id,
            "title": x.title,
            "teacher_id": x.teacher_id,
            "teacher_name": teacher_map.get(x.teacher_id),
            "materials": x.materials,
            "attachment_path": x.attachment_path,
            "direction": x.direction,
            "is_published": x.is_published,
        }
        for x in topics
    ]


@app.post("/students/{student_id}/project/teams/{team_id}/apply-topic")
def student_apply_project_topic(
    student_id: int,
    team_id: int,
    topic_id: int = Form(...),
    note: str = Form(default=""),
    attachment_path: str = Form(default=""),
    db: Session = Depends(get_db)
):
    student = db.query(models.User).filter(models.User.id == student_id, models.User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    team = db.query(models.ProjectTeam).filter(models.ProjectTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="队伍不存在")
    if team.leader_id != student_id:
        raise HTTPException(status_code=403, detail="仅队长可提交课题申请")

    current = _get_current_semester(db)
    if not current or team.semester_id != current.id:
        raise HTTPException(status_code=400, detail="仅支持当前学期申请")

    setting = _get_or_create_project_setting(db, current.id)
    if not _is_project_apply_open(setting):
        raise HTTPException(status_code=400, detail="当前不在综设组队开放时间")

    members_count = db.query(models.ProjectTeamMember).filter(models.ProjectTeamMember.team_id == team_id).count()
    if members_count < setting.min_team_size or members_count > setting.max_team_size:
        raise HTTPException(status_code=400, detail=f"队伍人数需在{setting.min_team_size}-{setting.max_team_size}人之间")

    topic = db.query(models.ProjectTopic).filter(models.ProjectTopic.id == topic_id).first()
    if not topic or not topic.is_published:
        raise HTTPException(status_code=400, detail="课题不存在或未发布")

    cfg = db.query(models.ProjectClassTopicConfig).filter(
        models.ProjectClassTopicConfig.class_id == team.class_id,
        models.ProjectClassTopicConfig.topic_id == topic.id
    ).first()
    if not cfg:
        raise HTTPException(status_code=400, detail="该班级未配置此课题")

    if topic.direction != team.direction:
        raise HTTPException(status_code=400, detail="队伍方向与课题方向不一致")

    team.topic_id = topic.id
    team.advisor_teacher_id = topic.teacher_id
    team.apply_note = note.strip() or None
    team.apply_attachment_path = attachment_path.strip() or None
    team.status = "pending"
    team.review_comment = None
    db.commit()

    return {"message": "课题申请已提交，等待老师审核"}


